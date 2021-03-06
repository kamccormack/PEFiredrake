# Kimmy McCormack
# Functions used by 'model_run_file.py' to solve coupled set of poroelastic equations in both 2D
# and 3D subduction zone domains
# Units are in meters and MPa

import sys
sys.path.append('/fenics/shared/local_lib')
import matplotlib.pyplot as plt
import numpy as np
import openpyxl as pyxl
import os
import scipy as sp
import scipy.interpolate
import scipy.io
import shutil
from firedrake import *
from firedrake.petsc import PETSc
from firedrake.matrix import MatrixBase
from scipy.interpolate import griddata
from tempfile import TemporaryFile
from pykrige.ok import OrdinaryKriging
import pykrige.kriging_tools as kt
import matplotlib.tri as tri
from IPython.display import HTML
import time


#Citations.print_at_exit()
##################### Optimization options ############################
parameters["reorder_meshes"] = False
parameters['form_compiler']['optimize'] = True
parameters["form_compiler"]["cpp_optimize"] = True
ffc_options = {"optimize": True, "eliminate_zeros": True,
               "precompute_basis_const": True, "precompute_ip_const": True}
parameters['allow_extrapolation'] = True


#################### Data interpolation functions ######################
def slabkriging(XYZslab, u0load, xvec, yvec, zvec, dof_slab, **kwargs):
	"""Interpolates surface z value data by kriging.

		INPUTS:
			[x,y,z] = coordinates of degrees of freedom onto which the z value is mapped
			[dof_surf] = the degrees of freedom of the surface of the mesh

			OrdinaryKriging function takes variogram parameters as:
				linear - [slope, nugget]
				power - [scale, exponent, nugget]
			`   gaussian - [sill, range, nugget]
				spherical - [sill, range, nugget]
				exponential - [sill, range, nugget]
				hole-effect - [sill, range, nugget]

		OUPUTS:
			z_surf - surface height on all mesh nodes as a numpy array
	"""
	
	
	
	x_slab, y_slab, z_slab = xvec[dof_slab], yvec[dof_slab], zvec[dof_slab]
	x_all, y_all = xvec[:], yvec[:]
	
	nonzero = np.where(np.abs(u0load[:,0]) >0.10)[0]
	skip = (slice(None, None, 5))
	
	uxslip_load, uyslip_load, uzslip_load = u0load[:,0][skip], \
	                                        u0load[:,1][skip], u0load[:,2][skip]
	xslip, yslip, zslip = XYZslab[:,0][skip], \
	                      XYZslab[:,1][skip], XYZslab[:,2][skip]
	
	# Interpolate data using pykrige
	OK_uxslip = OrdinaryKriging(xslip, yslip, uxslip_load, variogram_model='linear', \
	                            variogram_parameters=[2e-4, 1e-2])
	uxslip, ss_ux = OK_uxslip.execute('points', x_all, y_all)
	
	OK_uyslip = OrdinaryKriging(xslip, yslip, uyslip_load, variogram_model='linear', \
	                            variogram_parameters=[2e-4, 1e-2])
	uyslip, ss_uy = OK_uyslip.execute('points', x_all, y_all)
	
	OK_uzslip = OrdinaryKriging(xslip, yslip, uzslip_load, variogram_model='linear', \
	                            variogram_parameters=[2e-4, 1e-2])
	uzslip, ss_uz = OK_uzslip.execute('points', x_all, y_all)
	
	u03 = np.concatenate((uxslip.reshape((uxslip.shape[0],1)),
	                      uyslip.reshape((uxslip.shape[0],1)),
	                      uzslip.reshape((uxslip.shape[0],1))), axis=1)

	# np.save(var_path + 'uxslip', self.uxslip.compressed())
	# np.save(var_path + 'uyslip', self.uyslip.compressed())
	# np.save(var_path + 'uzslip', self.uzslip.compressed())
	
	return u03

def surfkriging(xvec, yvec, zvec, dof_surf, **kwargs):
	"""Interpolates surface z value data by kriging.

		INPUTS:
			[x,y,z] = coordinates of degrees of freedom onto which the z value is mapped
			[dof_surf] = the degrees of freedom of the surface of the mesh

			OrdinaryKriging function takes variogram parameters as:
				linear - [slope, nugget]
				power - [scale, exponent, nugget]
			`   gaussian - [sill, range, nugget]
				spherical - [sill, range, nugget]
				exponential - [sill, range, nugget]
				hole-effect - [sill, range, nugget]

		OUPUTS:
			z_surf - surface height on all mesh nodes as a numpy array
	"""

	x_surf, y_surf, z_surf = xvec[dof_surf], yvec[dof_surf], zvec[dof_surf]
	x_all, y_all = xvec[:], yvec[:]

	# PyKrige interpolation
	OK_surf = OrdinaryKriging(x_surf, y_surf, z_surf,
	                          variogram_model='gaussian', variogram_parameters=[5e5, 5e4, 1e2])
	z_surf, ss = OK_surf.execute('points', x_all, y_all)
	
	return z_surf.compressed()

def kappakriging(k_ocean, data_file, origin, theta, xvec, yvec, zvec, dof_surf, **kwargs):
	"""Interpolates surface permeability data by kriging. Take surface value to interpolate kappa with depth.

		INPUTS:
			[k_ocean] = log(ocean permebility) [m^2]
			[data_sheet] = excel sheet containing pointwise permeability data in (lon, lat, permaebility[m^2]) format.
					Default is to ignore the first row (containing headers).
			[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
			[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
					counter-clockwise from latlon to XY
			[log_kr] = log of residual permeabiltiy at depth (usually betwee -18 and -20) [m^2]
			[alpha_k] = decay constant of permeability as a function of depth (between 0.25 and 1.6)
			[coords_K] = coordinates of degrees of freedom onto which the permeability is mapped
			[dof_surf] = the degrees of freedom of the surface of the mesh

			OrdinaryKriging function takes variogram parameters as:
				linear - [slope, nugget]
				power - [scale, exponent, nugget]
			`   gaussian - [sill, range, nugget]
				spherical - [sill, range, nugget]
				exponential - [sill, range, nugget]
				hole-effect - [sill, range, nugget]

		OUPUTS:
			kappa [m^2] - the interpolated 3D permeabiltiy field in the form of a FEniCS Expression()
	"""

	x_surf, y_surf, z_surf = xvec[dof_surf], yvec[dof_surf], zvec[dof_surf]
	x_all, y_all = xvec[:], yvec[:]
	
	wb = pyxl.load_workbook(data_file)
	well_data = wb['well_data']

	x_wells, y_wells = latlon2XY(well_data, origin, theta)
	permeability = np.array(
		[[well_data.cell(row=i, column=5).value] for i in range(2, well_data.max_row + 1)]).reshape(
		well_data.max_row - 1, )
	logk = np.log10(permeability)

	# PyKrige interpolation
	# OK = OrdinaryKriging(x_wells, y_wells, logk, variogram_model = 'power', variogram_parameters = [1, 1, 2e4])
	OK = OrdinaryKriging(x_wells, y_wells, logk, variogram_model='exponential',
	                     variogram_parameters=[1.086, 1.28e4, 0.771])

	# k_surf, ss = OK.execute('points', x_surf, y_surf)
	k_surf, ss = OK.execute('points', x_all, y_all)

	#np.save("results/numpy_variables/k_surf", k_surf.compressed())
	# np.save("saved_variables_poroelastic/ss_krige", self.ss.compressed())
	# self.k_surf = np.load("saved_variables_poroelastic/k_surf_med.npy")
	# self.ss = np.load("saved_variables_poroelastic/ss_krige_med.npy")

	return k_surf.compressed()

def kappakriging2D(data_file, coords_K, dof_surf, **kwargs):
	"""Interpolates 2D permeability data by kriging from excel file

		INPUTS:
			[data_sheet] = excel sheet containing pointwise permeability data in (X, Y, permaebility[m^2]) format.
					Default is to ignore the first row (containing headers).
			[coords_K] = coordinates of degrees of freedom onto which the permeability is mapped
			[dof_surf] = the degrees of freedom of the surface of the mesh

			OrdinaryKriging function takes variogram parameters as:
				linear - [slope, nugget]
				power - [scale, exponent, nugget]
			`   gaussian - [sill, range, nugget]
				spherical - [sill, range, nugget]
				exponential - [sill, range, nugget]
				hole-effect - [sill, range, nugget]

		OUPUTS:
			kappa [m^2] - the interpolated 3D permeabiltiy field in the form of a FEniCS Expression()
	"""

	x_surf, y_surf = coords_K[dof_surf, 0], coords_K[dof_surf, 1]
	x_all, y_all = coords_K[:, 0], coords_K[:, 1]

	wb = pyxl.load_workbook(data_file)
	data = wb['kappa_data2D']

	x_kappa = 1e3 * np.array([[data.cell(row=i, column=1).value] for i in range(2, data.max_row + 1)]).reshape(
		data.max_row - 1, )
	y_kappa = 1e3 * np.array([[data.cell(row=i, column=2).value] for i in range(2, data.max_row + 1)]).reshape(
		data.max_row - 1, )
	permeability = np.array([[data.cell(row=i, column=3).value] for i in range(2, data.max_row + 1)]).reshape(
		data.max_row - 1, )
	logk = np.log10(permeability)

	# PyKrige interpolation

	# OK = OrdinaryKriging(x_wells, y_wells, logk, variogram_model = 'power', variogram_parameters = [1, 1, 2e4])
	OK = OrdinaryKriging(x_kappa, y_kappa, logk,
	                     variogram_model='gaussian', variogram_parameters=[4, 1e5, 0.5])

	# k_surf, ss = OK.execute('points', x_surf, y_surf)
	logk_all, ss = OK.execute('points', x_all, y_all)
	k_all = np.power(10, logk_all)

	# np.save("results/numpy_variables/k_all2D", k_all.compressed())
	# np.save("saved_variables_poroelastic/ss_krige", self.ss.compressed())
	# self.k_surf = np.load("saved_variables_poroelastic/k_all2D.npy")
	# self.ss = np.load("saved_variables_poroelastic/ss_krige_med.npy")

	# return k_all
	return k_all.compressed()

def topokriging(data, origin, theta, xvec, yvec, zvec, dof_surf, var_path, **kwargs):
	""" Maps surface topography to boundary pore pressure

	INPUTS:
		[data] = excel sheet containing topo data formatted as (lon, lat, elevation [meters])
		[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
		[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY
		[coords_Q] = coordinates of degrees of freedom onto which the topo pressure is mapped
		[dof_surf] = the degrees of freedom of the surface of the mesh
		[var_path] = the path of stored variables

	OUTPUTS:
		FEniCS Expression() for pore pressure along surface boundary
	"""

	x_surf, y_surf, z_surf = xvec[dof_surf], yvec[dof_surf], zvec[dof_surf]
	x_all, y_all = xvec[:], yvec[:]

	x_topo, y_topo = latlon2XY(data, origin, theta)
	topo_p = 0.7 * (9.81 / 1e3) * np.array(
		[[data.cell(row=i, column=3).value] for i in range(2, data.max_row + 1)]).reshape(data.max_row - 1, )

	# PyKrige interpolation
	OK_topo = OrdinaryKriging(x_topo, y_topo, topo_p,
	                          variogram_model='gaussian', variogram_parameters=[5e5, 5e4, 1e2])
	topo_p, sstopo = OK_topo.execute('points', x_surf, y_surf)

	np.save(var_path + 'topo_p', topo_p.compressed())
	# self.topo_p = np.load(var_path+'topo_p.npy')

	return topo_p.compressed()

def latlon2XY(data_sheet, origin, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	lon = np.array([[data_sheet.cell(row=i, column=1).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	lat = np.array([[data_sheet.cell(row=i, column=2).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )

	lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
	lat_in_km = (lat - origin[1]) * 111

	rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
	theta_new = np.arctan2(lat_in_km, lon_in_km) - theta

	X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)

	return 1e3 * X, 1e3 * Y

KExpression2D_cpp = '''
class KExpression2D_cpp : public Expression
{
public:

  KExpression2D_cpp() :
  Expression(),
  alphak(0),
  logkr(0)
  {
  }

    void eval(Array<double>& values, const Array<double>& x) const
    {
        Ksurf->eval(values, x);
        const double Ksurf_val = values[0];

        Zsurf->eval(values, x);
        const double Zsurf_val = values[0];

        values[0] = pow(10, (logkr + ((Ksurf_val - logkr) * pow(1 - (1e-3 * .15*(x[1] - Zsurf_val)), -alphak))));

    }

public:
    double alphak;
    double logkr;
    std::shared_ptr<const Function> Ksurf;
    std::shared_ptr<const Function> Zsurf;
};'''

KExpression3D_cpp = '''
class KExpression3D_cpp : public Expression
{
public:

  KExpression3D_cpp() :
  Expression(),
  alphak(0),
  logkr(0)
  {
  }

    void eval(Array<double>& values, const Array<double>& x) const
    {
        Ksurf->eval(values, x);
        const double Ksurf_val = values[0];

        Zsurf->eval(values, x);
        const double Zsurf_val = values[0];

        values[0] = pow(10, (logkr + ((Ksurf_val - logkr) * pow(1 - (1e-3 * (x[2] - Zsurf_val)), -alphak))));

    }

public:
    double alphak;
    double logkr;
    std::shared_ptr<const Function> Ksurf;
    std::shared_ptr<const Function> Zsurf;
};'''

class SlipExpression(Expression):
	""" Interpolates an arbitrary displacement field to the fault boundary of the mesh.

	INPUTS:
		[data] = an excel sheet with slip vectors recorded as
				(lon, lat, dip-slip [meters] , strike-slip [meters]) in the first four columns
		[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
		[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY
		[coords_V] = coordinates of degrees of freedom onto which the slip vectors are mapped
		[dof_slab] = the degrees of freedom of the fault boundary of the mesh
		[var_path] = the path of stored variables

	OUTPUTS:
		FEniCS Expression() containing the interpolated vector slip displacement on the fault boundary
	"""

	def __init__(self, data, origin, theta, coords_V, dof_slab, var_path, **kwargs):
		self.x_slab, self.y_slab, self.z_slab = coords_V[dof_slab, 0], coords_V[dof_slab, 1], coords_V[dof_slab, 2]

		X_slip, Y_slip = latlon2XY(data, origin, theta)

		Ux_slip = lon = np.array([[data.cell(row=i, column=3).value] for i in range(2, data.max_row + 1)]).reshape(
			data.max_row - 1, )
		Uy_slip = lon = np.array([[data.cell(row=i, column=4).value] for i in range(2, data.max_row + 1)]).reshape(
			data.max_row - 1, )

		# Interpolate data using pykrige
		OK_uxslip = OrdinaryKriging(X_slip, Y_slip, Ux_slip, variogram_model='linear', \
		                            variogram_parameters=[2e-4, 1e-2])
		self.uxslip, ss_ux = OK_uxslip.execute('points', self.x_slab, self.y_slab)

		OK_uyslip = OrdinaryKriging(X_slip, Y_slip, Uy_slip, variogram_model='linear', \
		                            variogram_parameters=[2e-4, 1e-2])
		self.uyslip, ss_ux = OK_uyslip.execute('points', self.x_slab, self.y_slab)

		np.save(var_path + 'uxslip', self.uxslip.compressed())
		np.save(var_path + 'uyslip', self.uyslip.compressed())

	# self.uxslip = np.load("saved_variables_poroelastic/uxslip_med.npy")
	# self.uyslip = np.load("saved_variables_poroelastic/uyslip_med.npy")

	def U_X_interp(self, X, Y):
		indx = (np.abs(np.sqrt(pow((self.x_slab[:] - X), 2) + pow((self.y_slab[:] - Y), 2)))).argmin()
		return self.uxslip[indx]

	def U_Y_interp(self, X, Y):
		indy = (np.abs(np.sqrt(pow((self.x_slab[:] - X), 2) + pow((self.y_slab[:] - Y), 2)))).argmin()
		return self.uyslip[indy]

	def eval(self, values, x, **kwargs):
		values[0] = self.U_X_interp(x[0], x[1])
		values[1] = self.U_Y_interp(x[0], x[1])
		values[2] = 0.0

	def value_shape(self):
		return (3,)
	
		
################### FEM solving functions #################################

class Poroelasticity2D:
	def __init__(self, data_file, ndim, mesh, boundaries, plot_figs, event, new_mesh, permeability, days, SSE_days,
	             sub_cycle_years, sigma_b, xcenter, SS_migrate, u0_EQ, u0_SSE, u0_sub, surface_k, ocean_k, B, loop,
	             loop_variable, percent_lith, dehydrate_depths, dehydrate_flux, xstick, depth_kappa, k_frack, crosssection, **kwargs):

		print("Running 2D model for a", event, " event for cross section at ", crosssection, " km")
		self.name_append = "cross_%s" % str(crosssection)

		self.start = time.time()
		self.data_file = data_file
		self.mesh = mesh
		self.boundaries = boundaries
		self.plot_figs = plot_figs
		self.event = event
		self.new_mesh = new_mesh
		self.permeability = permeability
		self.days = days
		self.SSE_days = SSE_days
		self.sub_cycle_years = sub_cycle_years
		self.sigma_b = sigma_b
		self.xcenter = xcenter
		self.SS_migrate = SS_migrate
		self.u0_EQ = u0_EQ
		self.u0_SSE = u0_SSE
		self.u0_sub = u0_sub
		self.surface_k = surface_k
		self.ocean_k = ocean_k
		self.loop = loop
		self.loop_variable = loop_variable
		self.percent_lith = percent_lith
		self.dehydrate_depths = dehydrate_depths
		self.dehydrate_flux = dehydrate_flux
		self.xstick = xstick
		self.k_depth = depth_kappa
		self.k_increase = k_frack
		self.crosssection = crosssection*1e3

		# self.solver = KrylovSolver('bicgstab', 'ilu')
		# self.solver = LUSolver('umfpack')
		self.solver = LinearSolver('mumps')
		self.prm = self.solver.parameters
		self.prm['reuse_factorization'] = True  # Saves Factorization of LHS

		self.V = VectorElement('Lagrange', self.mesh.ufl_cell(), 2)
		self.Q = FiniteElement('Lagrange', self.mesh.ufl_cell(), 1)
		self.ME = FunctionSpace(self.mesh, MixedElement([self.Q, self.V]))
		self.Vspace = FunctionSpace(self.mesh, self.V)
		self.Qspace = FunctionSpace(self.mesh, self.Q)
		self.Q_gspace = VectorFunctionSpace(self.mesh, "CG", 1)
		self.Wspace = TensorFunctionSpace(self.mesh, "Lagrange", 1)
		
		self.Q2 = FiniteElement('Lagrange', self.mesh.ufl_cell(), 2)
		self.Q2space = FunctionSpace(self.mesh, self.Q2)

		self.Kelement = FiniteElement('CG', self.mesh.ufl_cell(), 1)
		self.Kspace = FunctionSpace(self.mesh, self.Kelement)

		(self.p, self.u) = TrialFunctions(self.ME)
		(self.q, self.v) = TestFunctions(self.ME)
		self.ds = Measure("ds", domain=self.mesh, subdomain_data=self.boundaries)  #
		self.n = FacetNormal(self.mesh)  # normal vector
		self.dim = self.ME.dim()
		self.pdim = self.ME.sub(0).dim()
		self.udim = self.ME.sub(1).dim()
		self.geom_dim = ndim + 1

		self.results_path = 'results/numpy_results/'
		self.var_path = 'results/numpy_variables/'
		self.paraviewpath = 'results/paraview/'
		self.inverse_results = '/home/fenics/shared/inverse_codes/results_deterministic/numpy_results/'
		self.inverse_var = '/home/fenics/shared/inverse_codes/results_deterministic/numpy_variables/'
		
		if not os.path.exists(self.paraviewpath):
			os.makedirs(self.paraviewpath)

		self.extract_coords()
		print("Coordinates extracted", (time.time() - self.start) / 60.0, "minutes")

		self.k_interpolation()
		print("Permeability interpolated", (time.time() - self.start) / 60.0, "minutes")
		
		self.slip_interpolation()
		print("Slip interpolated", (time.time() - self.start) / 60.0, "minutes")
		
		self.time_stepping()

		######################### POROELASTIC PARAMETERS #############################
		self.delta = 1e-6  # Dirichlet control regularization on bottom boundary
		self.muf = Constant('1e-9')  # pore fluid viscosity [MPa s]
		self.B = B  # Skempton's coefficient  ...Should this be constant?
		self.nuu = .4  # undrained poissons ratio
		self.nu = 0.25  # Poisson's ratio
		# Young's modulus [MPa]. Fit from Deshon paper (2006?)
		# self.E = Expression('(5.7972e-10*(-(pow(x[1],3)))) - (7.25283e-5*(-(pow(x[1],2)))) + (3.8486e-6*(-x[1])) + 6.94e4', degree = 1)
		self.E = Expression('500 * pow(-x[1], 0.516) + 2e4', degree=1)
		#self.E = Expression('347 * pow(-x[1], 0.516) + 5.324e4', degree=1)
		self.mu = Expression('E / (2.0*(1.0 + nu))', E=self.E, nu=self.nu, degree=1)  # shear modulus
		self.lmbda = Expression('E*nu / ((1.0 + nu)*(1.0 - 2.0*nu))', E=self.E, nu=self.nu,
		                        degree=1)  # Lame's parameter
		self.alpha = 3 * (self.nuu - self.nu) / (self.B * (1 - 2 * self.nu) * (1 + self.nuu))  # Biot's coefficient
		self.Se = Expression('(9*(nuu-nu)*(1-(2*nuu)))/(2*mu*pow(B,2)*(1-2*nu)*pow((1+nuu),2))', nuu=self.nuu,
		                     nu=self.nu, mu=self.mu, B=self.B, degree=1)  # mass specific storage
		self.kstick = 3e-5
		self.stick = Expression('0.8*(1/(1+exp(-kstick*(xstick-x[0]))))', xstick=self.xstick, kstick=self.kstick,
		                        degree=1)
		self.d = self.u.geometric_dimension()  # number of space dimensions
		self.I = Identity(self.d)
		self.T = (self.I - outer(self.n, self.n))  # tangent operator for boundary condition
		self.year = 60.0 * 60.0 * 24.0 * 365.0
		self.day = 60.0 * 60.0 * 24.0

		self.rhof = Constant('1e-3')  # pore fluid density [kg *1e-6]
		self.rho_rock = Constant('2.7e-3')  # rock density
		self.g = Constant('9.81')  # gravity

		######################### VERTICAL AND HORIZONTAL STRESS  #############################

		self.p_hydrostatic = Expression('-rho_f * g * x[1]', rho_f=self.rhof, g=self.g, degree=1)
		self.sig_v = Expression('-rho_rock * g * x[1]', rho_rock=self.rho_rock, g=self.g, degree=1)
		self.p_f = Expression('(1-per_lith)*(sigv - ph)', sigv=self.sig_v,
		                      ph=self.p_hydrostatic, per_lith=self.percent_lith, degree=1)
		self.p_fh = Expression('(sigv - ph) +.1', sigv=self.sig_v,
		                       ph=self.p_hydrostatic, per_lith=self.percent_lith, degree=1)
		self.sig_min = Expression('0', degree=1)
		# self.del_p = Expression('rhof * 9.81 * x[1]', rhof = self.rhof, degree = 1)
		self.op = Constant('0.8')
		self.dof_frack_old = np.empty([1, ])
		self.dof_unfrack_old = np.empty([1, ])

		self.phydro = Function(self.Qspace)
		self.phydro.interpolate(self.p_fh)
		self.p_frack = Function(self.Qspace)
		self.p_frack.interpolate(self.p_f)
		self.dof_shallow = np.where(self.coords_Q[:, 1] >= -4.0e3)[0]
		self.p_frack.vector()[self.dof_shallow] = self.phydro.vector()[self.dof_shallow]

		################### INITIAL AND BOUNDARY CONDITIONS ######################
		####### Initial conditions: (('p0','ux0','uy0')) #########

		# self.p0 = Expression('op*(sv-ph)', op = self.op, sv=self.sig_v, ph=self.p_hydrostatic, degree=1)
		# self.p0 = Expression('ph', ph=self.p_hydrostatic, degree=1)
		# self.w_init = Expression(('p0', '0', '0'), p0 = self.p0, degree=1)
		self.w_init = Expression(('0', '0', '0'), degree=1)

		self.w0 = Function(self.ME)
		self.w0.interpolate(self.w_init)
		self.ub0 = Expression(('0', '0'), degree=1)
		self.ub1 = Expression(('1', '1'), degree=1)
		self.b0 = Constant('0')
		self.b1 = Constant('1.0')
		self.bc1 = DirichletBC(self.ME.sub(1), self.ub0,  2)  # top/surface
		self.bc2 = DirichletBC(self.ME.sub(1), self.ub0,  3)  # bottom/fault boundary
		self.bc3 = DirichletBC(self.ME.sub(1).sub(0), self.b0, self.boundaries,
		                       4)  # right/ back side (no slip condition)
		# self.bc3 = DirichletBC(self.ME.sub(1), self.ub0,  4)  # right/ back side (no slip condition)
		self.bc4 = DirichletBC(self.ME.sub(1), self.ub0,  5)  # mantle (no slip condition)
		# self.bc4 = DirichletBC(self.ME.sub(1), self.ub0,  5)  # right/ back side (no slip condition)
		self.bc5 = DirichletBC(self.ME.sub(1), self.ub0,  1)  # ocean surface
		self.bcp1 = DirichletBC(self.ME.sub(0), self.b0,  1)  # ocean surface (free flow condition)
		self.bcp2 = DirichletBC(self.ME.sub(0), self.b0,  2)  # land surface (free flow condition)

		if self.event == 'topo':
			self.topo_flow_interp()
			self.bcp2 = DirichletBC(self.ME.sub(0), self.p_topo,  2)  # land surface
			self.bcs = [self.bcp1, self.bcp2]  # These are the BC's that are actually applied
		else:
			self.bcs = [self.bc3, self.bc4, self.bcp1]  # These are the BC's that are actually applied
			# self.bcs = [self.bc3, self.bc4, self.bc1, self.bc5]

		self.slab = 3  # the slab is labeled as boundary 3 for this mesh, but not always the case

		################### SET UP TO SAVE SOLUTION  ######################
		self.Sol_all = np.empty((self.dim, self.nsteps))  # matrix to save entire solution
		self.Sol_surf = np.empty((self.surface_dofs.shape[0], self.nsteps))  # matrix to save surface solution
		self.sea_flux = np.empty((self.ocean_dof_pgrad.shape[0], self.nsteps))  # ocean bottom solution
		self.flow_velocity = np.empty((self.pdim * ndim, self.nsteps))  # ocean bottom solution
		self.Psi = np.empty((self.size_p, self.nsteps))  # ocean bottom solution
		self.vol_strain = np.empty((self.size_p, self.nsteps))  # ocean bottom solution
		self.sea_flux_total = np.empty(self.nsteps)  # matrix to save surface post-seismic solution
		self.esigv_slab = np.empty((self.dof_slabQ.shape[0], self.nsteps))

		if loop == 'yes':
			if self.loop_variable == 'kappa':
				self.loop_name = "k" + str(-self.surface_k)
				pfilename = self.paraviewpath + "pressure2D_loop_k%s.pvd" % str(-self.surface_k)
				p_velfilename = self.paraviewpath + "pressure_vel2D_loop_k%s.pvd" % str(-self.surface_k)
				ufilename = self.paraviewpath + "def2D_loop_k%s.pvd" % str(-self.surface_k)
				#sigfilename = self.paraviewpath + "stress2D_k%s.pvd" % str(-self.surface_k)
				#effsigfilename = self.paraviewpath + "effstress2D_k%s.pvd" % str(-self.surface_k)
				#eigsigfilename = self.paraviewpath + "eigstress2D_k%s.pvd" % str(-self.surface_k)
			elif self.loop_variable == 'sigma':
				self.loop_name = 'sig' + str(self.sigma_b)[:2]
				pfilename = self.paraviewpath + "pressure2D_loop_%s.pvd" % str(self.sigma_b)[:2]
				p_velfilename = self.paraviewpath + "pressure_vel2D_loop_%s.pvd" % str(self.sigma_b)[:2]
				ufilename = self.paraviewpath + "def2D_loop_%s.pvd" % str(self.sigma_b)[:2]
				sigfilename = self.paraviewpath + "stress2D_%s.pvd" % str(self.sigma_b)[:2]
				effsigfilename = self.paraviewpath + "effstress2D_%s.pvd" % str(self.sigma_b)[:2]
				eigsigfilename = self.paraviewpath + "eigstress2D_%s.pvd" % str(self.sigma_b)[:2]
			elif self.loop_variable == 'B':
				self.loop_name = 'B' + str(-self.B)
				pfilename = self.paraviewpath + "pressure2D_loop_B%s.pvd" % str(-self.B)
				p_velfilename = self.paraviewpath + "pressure_vel2D_loop_B%s.pvd" % str(-self.B)
				ufilename = self.paraviewpath + "def2D_loop_k%s.pvd" % str(-self.B)
				sigfilename = self.paraviewpath + "stress2D_B%s.pvd" % str(-self.B)
				effsigfilename = self.paraviewpath + "effstress2D_B%s.pvd" % str(-self.B)
				eigsigfilename = self.paraviewpath + "eigstress2D_B%s.pvd" % str(-self.B)
			elif self.loop_variable == 'cross':
				self.loop_name = "cross" + str(self.crosssection)[0:3]
				pfilename = self.paraviewpath + "pressure2D_loop_k%s.pvd" % str(self.crosssection)[0:3]
				p_velfilename = self.paraviewpath + "pressure_vel2D_loop_k%s.pvd" % str(self.crosssection)[0:3]
				ufilename = self.paraviewpath + "def2D_loop_k%s.pvd" % str(self.crosssection)[0:3]
		else:
			pfilename = self.paraviewpath + "pressure2D" + self.name_append + ".pvd"
			# pnormfilename = self.paraviewpath + "pressure_norm2D"+self.name_append+".pvd"
			p_velfilename = self.paraviewpath + "pressure_vel2D" + self.name_append + ".pvd"
			ufilename = self.paraviewpath + "def2D" + self.name_append + ".pvd"
			# sigfilename = self.paraviewpath + "stress2D" + self.name_append + ".pvd"
			# effsigfilename = self.paraviewpath + "effstress2D" + self.name_append + ".pvd"
			# eigsigfilename = self.paraviewpath + "eigstress2D" + self.name_append + ".pvd"
			# eigeffsigfilename = self.paraviewpath + "eigeffstress2D" + self.name_append + ".pvd"
			# frackingfilename = self.paraviewpath + "fracking2D" + self.name_append + ".pvd"

		self.pfile = File(pfilename)
		# self.pnormfile = File(pnormfilename)
		self.p_velfile = File(p_velfilename)
		self.ufile = File(ufilename)
		# self.sigfile = File(sigfilename)
		# self.effsigfile = File(effsigfilename)
		# self.eigsigfile = File(eigsigfilename)
		# self.eigeffsigfile = File(eigeffsigfilename)
		# self.frackingfile = File(frackingfilename)

		################### SET UP, RUN, SOLVE and SAVE  ######################

		self.plot_k()
		print("Permeability plotted", (time.time() - self.start) / 60.0, "minutes")
		self.plot_mu()
		print("Shear modulus plotted", (time.time() - self.start) / 60.0, "minutes")
		self.assemble_system()
		print("System assembled", (time.time() - self.start) / 60.0, "minutes")
		# self.plot_initial_cond()
		self.solve_system()
		self.save_solution()

	def time_stepping(self):
		if self.event == 'EQ':  # 15 min for 1/2 day, 4 hrs for 10 days, 12 hours for 30 days, 2 days for 2 months, per month for 10 years, per year for 50 years
			self.dtt = [60, 240, 720, 288e1, 2 * 432e2]  # , 5256e2]  # in minutes
			self.ntime = [24, 60, 60, 30, 60]  # , 50]
			#self.dtt = [15]  # in minutes
			#self.ntime = [1]
		elif self.event == 'SSE':  # 12 hrs for SSE days, 12 hrs for total time-SSE_days, 2 months for 10 years
			self.dtt = [288e1 / 2, 432e2 * 2]  # 4 days, 2 months
			self.ntime = [self.SSE_days / 1, (self.days - self.SSE_days) / 60]  # SSE_days, 10 years
			# self.dtt = [288e1*2, 432e2*12]  # 6 months
			# self.ntime = [1, 10] # 10 years
			# self.dtt = [720]  # in minutes
			# self.ntime = [self.SSE_days * 2]
		elif self.event == 'sub_EQ':  # 1/year for 50 years followed by 'EQ' scheme
			self.dtt = [60 * 24 * 365, 15, 240, 288e1, 432e2, 60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years, 96, 120, 15, 122, 40]
		elif self.event == 'sub':  # 1/year for 50 years
			self.dtt = [60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years]
		elif self.event == 'topo':  # 1/year for 100 years
			self.dtt = [60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years]
		self.dtt = [i * 60 for i in self.dtt]  # in seconds
		self.dtt_repeated = np.repeat(self.dtt, self.ntime)
		self.dtt_comsum = np.cumsum(self.dtt_repeated)
		self.dt = self.dtt_comsum[0]
		self.nsteps = self.dtt_comsum.size
	
	def slip_interpolation(self):
		tic = time.time()
		## import inverse solution and mesh ##
		mesh_size_inv = 'med_inv'  # fine, med, coarse
		path = "/home/fenics/shared/meshes"
		invmesh = Mesh(path + '/CR3D_' + mesh_size_inv + '.xml')
		
		# invmesh = Mesh()
		input_file = HDF5File(self.mesh.mpi_comm(), self.inverse_results + "u0slab.h5", "r")
		# input_file.read(invmesh, 'mesh')
		Vh = VectorFunctionSpace(invmesh, 'Lagrange', 1)
		u0load = Function(Vh)
		ufwdload = Function(Vh)
		input_file.read(u0load, 'slip')
		input_file.close()
		
		## import dofs of slab boundary from inverse solution
		dofslab_inv = np.load(self.inverse_var + 'slab_dofs_u_3D.npy')
		xyz_slab = np.load(self.inverse_var + 'XYZ_slab.npy')
		xslab_ind = xyz_slab.shape[0]/3
		xslab, yslab, zslab = xyz_slab[0:xslab_ind, 0], xyz_slab[0:xslab_ind, 1], xyz_slab[0:xslab_ind, 2]
		Ux_slab = u0load.vector()[dofslab_inv][0:xslab_ind]
		Uz_slab = u0load.vector()[dofslab_inv][2*xslab_ind::]
		cross_ind = np.where(np.sqrt((yslab - self.crosssection)**2) <= 1e3)[0]
		xslab_cross, zslab_cross = xslab[cross_ind], zslab[cross_ind]
		Ux_cross, Uz_cross = Ux_slab[cross_ind], Uz_slab[cross_ind]
		
		
		x_Q, y_Q = self.coords_Q[:, 0], self.coords_Q[:, 1]
		bctest = DirichletBC(self.Qspace, (3),  3)  # slab surface
		ptest = Function(self.Qspace)
		bctest.apply(ptest.vector())
		slab_dofsQ = np.where(ptest.vector() == 3)[0]
		
		X_slab, Z_slab = x_Q[slab_dofsQ], y_Q[slab_dofsQ]
		
		print("cross ind", cross_ind.shape)

		# Interpolate data using pykrige
		OK_uxslip = OrdinaryKriging(xslab_cross, zslab_cross, Ux_cross, variogram_model='linear',
		                            variogram_parameters=[2e-4, 1e-2])
		ux, ss_ux = OK_uxslip.execute('points', X_slab, Z_slab)

		OK_uyslip = OrdinaryKriging(xslab_cross, zslab_cross, Uz_cross, variogram_model='linear',
		                            variogram_parameters=[2e-4, 1e-2])
		uz, ss_uz = OK_uyslip.execute('points', X_slab, Z_slab)

		
		# u0loadslab = Function(Vh)
		# # u0loadslab.vector()[:] = 0.0
		# u0loadslab.vector()[dofslab_inv] = u0load.vector()[dofslab_inv]
		#
		# self.u0slabV = Function(self.V1space)
		# # self.u0slabV.assign(project(u0loadslab, self.V1space, solver_type='mumps'))
		# self.u0slabV.assign(project(u0loadslab, self.V1space, solver_type='gmres'))
		#
		# uxdim = self.u0slabV.vector()[:].array().shape[0] / 3
		#
		# ux = self.u0slabV.vector().array()[0:uxdim]
		# uy = self.u0slabV.vector().array()[uxdim:2 * uxdim]
		# uz = self.u0slabV.vector().array()[2 * uxdim:3 * uxdim]
		
		u0x_func = Function(self.Qspace)
		u0z_func = Function(self.Qspace)
		
		u0x_func.vector()[slab_dofsQ] = ux
		u0z_func.vector()[slab_dofsQ] = uz
		
		self.u0slab = Expression(('ux', 'uz'), ux=u0x_func, uz=u0z_func, degree=1)

	def k_interpolation(self):
		log_kr = -18.
		alpha_k = 0.9

		self.coords_K = self.Kspace.tabulate_dof_coordinates().reshape((self.Kspace.dim(), -1))
		bctest1 = DirichletBC(self.Kspace, (1),  1)  # ocean surface
		bctest2 = DirichletBC(self.Kspace, (1),  2)  # land surface
		ktest = Function(self.Kspace)
		bctest1.apply(ktest.vector())
		bctest2.apply(ktest.vector())
		self.dof_surfK = np.where(ktest.vector() == 1)[0]

		if self.permeability == 'mapped':

			kappa_2D = kappakriging2D(self.data_file, self.coords_K, self.dof_surfK)
			self.k_2D_func = Function(self.Kspace)
			self.k_2D_func.vector()[:] = 1e-18
			self.kappa = Expression('kappa', kappa=self.k_2D_func, degree=1)
			# self.kappa = Expression('1e-18', degree=1)

		elif self.permeability == 'constant':
			# self.kappa = Expression('pow(10, kappa)', kappa=self.surface_k, degree=1)
			# self.kappa = KExpression2D(self.surface_k, self.ocean_k, log_kr, alpha_k, self.coords_K, self.dof_surfK, degree=1)
			kappa_surf = self.surface_k
			self.k_surf_func = Function(self.Kspace)
			self.z_surf_func = Function(self.Kspace)
			self.k_surf_func.vector()[:] = kappa_surf
			#self.z_surf_func.vector()[self.dof_surfK] = self.coords_K[self.dof_surfK, 1]
			self.z_surf_func.vector()[:] = 0.


			self.kappa = Expression(KExpression2D_cpp, degree=1)
			self.kappa.alphak = alpha_k
			self.kappa.logkr = log_kr
			self.kappa.Ksurf = self.k_surf_func
			self.kappa.Zsurf = self.z_surf_func

			#self.k_2D_func = Function(self.Kspace)
			#self.k_2D_func.interpolate(self.kappa)

	def k_frack(self, p):

		self.p_fracking = p - self.p_frack
		self.frack = Function(self.Qspace)
		self.frack.assign(project(self.p_fracking, self.Qspace, solver_type='mumps'))

		self.dof_frack = np.where(self.frack.vector() > 0.01)[0]
		self.dof_unfrack = np.where(self.frack.vector()[self.dof_frack_old] <= 0.0)[0]

		self.dof_frack_new = np.setdiff1d(self.dof_frack, self.dof_frack_old)
		self.dof_unfrack_new = np.setdiff1d(self.dof_unfrack, self.dof_unfrack_old)

		print("frack new: ", self.dof_frack_new[:])
		print("seal new: ", self.dof_unfrack_new[:])

		self.k_2D_func.vector()[self.dof_frack_new] *= self.k_increase
		self.k_2D_func.vector()[self.dof_frack_old[self.dof_unfrack_new]] *= 10 / self.k_increase
		self.kappa = Expression('kappa', kappa=self.k_2D_func, degree=1)

		self.kappaplot.interpolate(Expression('log10(kappa)', kappa=self.kappa, degree=1))
		self.kappafile << self.kappaplot

		self.dof_frack_old = np.concatenate([self.dof_frack_old, self.dof_frack_new])
		self.dof_unfrack_old = np.concatenate((self.dof_unfrack_old, self.dof_unfrack_new))

	def kappa_update(self, dt):
		self.a_K = assemble(
			((self.kappa) / self.muf) * inner(nabla_grad(self.p), nabla_grad(self.q)) * dx)  # Stiffness matrix
		self.A = self.a_Mass + dt * self.a_K + self.alpha * self.a_Div + self.alpha * self.a_Grad \
		         + self.a_E + self.a_Boundary  # Left hand side (LHS)

	def topo_flow_interp(self):  # debug

		wb = pyxl.load_workbook(self.data_file)
		topo_data = wb['topo_data']
		x_topo = np.array([[topo_data.cell(row=i, column=1).value] for i in range(2, topo_data.max_row + 1)]).reshape(
			topo_data.max_row - 1, )
		p_topo = 0.7 * (9.81 / 1e3) * np.array(
			[[topo_data.cell(row=i, column=2).value] for i in range(2, topo_data.max_row + 1)]).reshape(
			topo_data.max_row - 1, )
		self.coords_K = self.Kspace.tabulate_dof_coordinates().reshape((self.Kspace.dim(), -1))
		bctest1 = DirichletBC(self.Kspace, (1),  1)  # ocean surface
		bctest2 = DirichletBC(self.Kspace, (1),  2)  # land surface
		qtest = Function(self.Kspace)
		bctest1.apply(qtest.vector())
		bctest2.apply(qtest.vector())
		self.dof_surfK = np.where(qtest.vector() == 1)[0]

		topo_surf_p = griddata(x_topo, p_topo, self.coords_K[self.dof_surfK, 0], method='linear')

		p_surf_func = Function(self.Kspace)
		p_surf_func.vector()[self.dof_surfK] = topo_p_surf
		self.p_topo = Expression(('ptopo'), ptopo=p_surf_func, degree=1)

	def plot_k(self):
		self.kappafile = File(self.paraviewpath + 'kappa_2D.pvd')
		self.kappaplot = Function(self.Qspace)
		self.kappaplot.interpolate(Expression('log10(kappa)', kappa=self.kappa, degree=1))
		print(self.kappaplot.vector().min(), self.kappaplot.vector().max())
		self.kappafile << self.kappaplot

	def plot_mu(self):
		self.Gfile = File(self.paraviewpath + 'Shearmod_2D.pvd')
		muplot = Function(self.Qspace)
		muplot.interpolate(Expression('mu', mu=self.Se, degree=1))
		# print(muplot.vector().min(), muplot.vector().max())
		self.Gfile << muplot

	def strain(self, w):  # strain = 1/2 (grad u + grad u^T)
		return sym(nabla_grad(w))

	def sigma(self, w):  # stress = 2 mu strain + lambda tr(strain) I
		return 2.0 * self.mu * self.strain(w) + self.lmbda * div(w) * self.I

	def assemble_system(self):
		"""NO TOUCHY... that's how you get ants!"""
		self.a_Mass = assemble(self.Se * self.p * self.q * dx)  # Mass matrix
		self.a_K = assemble(
			((self.kappa) / self.muf) * inner(nabla_grad(self.p), nabla_grad(self.q)) * dx)  # Stiffness matrix
		self.a_Div = assemble(nabla_div(self.u) * self.q * dx)  # Divergence matrix
		self.a_Grad = assemble(-self.p * (nabla_div(self.v)) * dx)  # Gradient
		self.a_E = assemble(inner(self.sigma(self.u), nabla_grad(self.v)) * dx)  # Elasticity
		self.a_Boundary = assemble((1 / self.delta) * dot(self.u, self.n) * dot(self.v, self.n) * self.ds(self.slab)
		                           + (1 / self.delta) * inner(self.T * self.u, self.T * self.v) * self.ds(self.slab))

		self.A = self.a_Mass + self.dt * self.a_K + self.alpha * self.a_Div + self.alpha * self.a_Grad \
		         + self.a_E + self.a_Boundary  # Left hand side (LHS)
		self.L = self.a_Mass + self.alpha * self.a_Div  # RHS

	def extract_coords(self):
		if not os.path.exists(self.var_path):
			os.makedirs(self.var_path)
		if self.new_mesh == 'yes':  # create surface/GPS indices and save them

			self.coordinates = self.ME.tabulate_dof_coordinates().reshape((self.dim, -1))
			self.x_all, self.y_all = self.coordinates[:, 0], self.coordinates[:, 1]

			self.coords_Q = self.Qspace.tabulate_dof_coordinates().reshape((self.pdim, -1))
			self.coords_V = self.Vspace.tabulate_dof_coordinates().reshape((self.udim, -1))

			bctest0 = DirichletBC(self.ME, (1, 1, 1),  1)  # ocean surface
			bctest1 = DirichletBC(self.ME, (2, 2, 2),  2)  # top surface
			bctest2 = DirichletBC(self.ME, (3, 3, 3),  3)  # slab surface

			ptest = Function(self.ME)
			bctest0.apply(ptest.vector())
			bctest1.apply(ptest.vector())
			bctest2.apply(ptest.vector())
			self.ocean_dofs = np.where(ptest.vector() == 1)[0]
			self.surface_dofs = np.where(ptest.vector() == 2)[0]
			self.slab_dofs = np.where(ptest.vector() == 3)[0]

			self.size_p = self.pdim
			self.size_u = self.dim - self.size_p

			self.ocean_dof_p = self.ocean_dofs[np.where(self.ocean_dofs < self.size_p)]
			self.ocean_dof_u = self.ocean_dofs[np.where(self.ocean_dofs >= self.size_p)] - self.size_p
			self.ocean_dof_pgrad = np.hstack((self.ocean_dof_p, (self.ocean_dof_p + self.ocean_dof_p.shape[0])))
			self.surface_dof_p = self.surface_dofs[np.where(self.surface_dofs < self.size_p)]
			self.surface_dof_u = self.surface_dofs[np.where(self.surface_dofs >= self.size_p)] - self.size_p
			self.surface_dof_pgrad = np.hstack((self.surface_dof_p, (self.surface_dof_p + self.surface_dof_p.shape[0])))

			bctestq = DirichletBC(self.Qspace, (1),  3)  # slab surface
			qtest = Function(self.Qspace)
			bctestq.apply(qtest.vector())
			self.dof_slabQ = np.where(qtest.vector() == 1)[0]

			np.save(self.var_path + "dof_slab", self.slab_dofs)
			np.save(self.var_path + "x_all_2D", self.x_all)
			np.save(self.var_path + "y_all_2D", self.y_all)
			np.save(self.var_path + "surface_dofs_2D", self.surface_dofs)
			np.save(self.var_path + "ocean_dofs_2D", self.ocean_dofs)
			# np.save(self.var_path+"indices_surf_CR_2D", self.indices_surf)
			# np.save(self.var_path+"size_surf_CR_2D", self.size_surf)
			np.save(self.var_path + "size_p_CR_2D", self.size_p)
		elif self.new_mesh == 'no':  # load saved indices/variables
			# self.x_all = np.load(self.var_path+"x_all_2D.npy")
			# self.y_all = np.load(self.var_path+"y_all_2D.npy")
			self.x_surf = np.load(self.var_path + "x_surf_2D.npy")
			self.y_surf = np.load(self.var_path + "y_surf_2D.npy")
			self.indices_surf = np.load(self.var_path + "indices_surf_CR_2D.npy")
			self.size_surf = np.load(self.var_path + "size_surf_CR_2D.npy")

	def init_boundary_slip(self, dt):
		if self.event == 'SSE':
			self.u_addSS = self.u0_SSE
			self.u_addsub = self.u0_sub
			self.u0_SSE = self.u_addSS * (dt / self.day)
			self.u0_sub = self.u_addsub * (dt / self.year)

		elif self.event == 'sub':  # Amount of slip per day (at center of event)
			self.u_add = self.u0_sub * (dt / self.year)
		elif self.event == 'sub_EQ':  # Amount of slip per day (at center of event)
			self.u_add = self.u0_sub * (dt / self.year)

	def point_source(self, dt):

		dehydrate_depths = self.dehydrate_depths
		dehydrate_flux = dt * self.dehydrate_flux
		# print('dehydrate flux: ', dehydrate_flux)

		for i in range(0, len(dehydrate_depths)):
			dof = np.where(self.coords_Q[:, 1] == self.coords_Q[self.dof_slabQ[np.abs(self.coords_Q[self.dof_slabQ, 1]
			                                                                          - dehydrate_depths[
				                                                                          i]).argmin()], 1])[0]
			PointSource(self.Qspace, Point(self.coords_Q[dof, 0][0], self.coords_Q[dof, 1][0]),
			            dehydrate_flux[i]).apply(self.b)

	def plot_initial_cond(self):  # plot the initial conditions
		plot(self.w0.sub(1), mode="glyphs", key="def")  # , rescale = False)
		plot(self.w0.sub(0), mode='color', key="pressure")  # , rescale = False)

	def streamfunction(self):
		"""Stream function for a given general 2D velocity field.
		The boundary conditions are weakly imposed through the term

			inner(q, grad(psi)*n)*ds,

		where grad(psi) = [-v, u] is set on all boundaries.
		This should work for any collection of boundaries:
		walls, inlets, outlets etc.
		"""
		vel = self.q_velocity
		qtest = TestFunction(self.Qspace)
		psi = TrialFunction(self.Qspace)
		a = inner(grad(qtest), grad(psi)) * dx
		# L   = dot(qtest, u[1].dx(0) - u[0].dx(1))*dx
		L = dot(qtest, curl(vel)) * dx
		bcu = []
		L = L + qtest * (self.n[1] * vel[0] - self.n[0] * vel[1]) * self.ds

		# Compute solution
		psi = Function(self.Qspace)
		A = assemble(a)
		b = assemble(L)
		normalize(b)  # Because we only have Neumann conditions
		[bc.apply(A, b) for bc in bcu]
		solve(A, psi.vector(), b)
		normalize(psi.vector())

		return psi

	def solve_system(self):
		count = 0
		if self.event == 'EQ':  # For an EQ, this can be outside the time loop
			#u0slab = Expression(('u0*exp(-(pow((x[0] - xcenter),2)/(pow(sigma,2))))', '0'), u0=self.u0_EQ,
			#                    xcenter=self.xcenter, sigma=self.sigma_b, degree=1)
			# u0slab = Expression(('0', '0'),degree = 1)
			a_Boundary0 = assemble((1 / self.delta) * inner(self.u0slab, self.T * self.v) * self.ds(
				self.slab))

		######## TIME LOOP ########
		for i in range(self.nsteps):

			if i == 0:
				dt = self.dtt_comsum[i]
			else:
				dt = self.dtt_comsum[i] - self.dtt_comsum[i - 1]
			if i > 2:
				dt_old = self.dtt_comsum[i - 1] - self.dtt_comsum[i - 2]
			if i > 2 and dt != dt_old:
				self.A = self.a_Mass + dt * self.a_K + self.alpha * self.a_Div + self.alpha * self.a_Grad \
				         + self.a_E + self.a_Boundary  # Left hand side (LHS)

			# if i > 0:
			# 	if self.event == 'SSE' and dt * count <= self.SSE_days * self.day:
			# 		self.u0_SSE += self.u_addSS * (dt / self.day)
			# 		self.xcenter += self.SS_migrate * (dt / self.day)
			# 		self.u0_sub += self.u_addsub * (dt / self.year)
			# 	if self.event == 'SSE' and dt * count > self.SSE_days * self.day:
			# 		self.u0_sub += self.u_addsub * (dt / self.year)
			# 	elif self.event == 'sub':
			# 		self.u0_sub += self.u_add
			# 	elif self.event == 'sub_EQ':
			# 		self.u0_sub += self.u_add * (dt / self.year)
			#else:
			#	self.init_boundary_slip(dt)

			# # Slip boundary Expression
			# if self.event == 'SSE':
			# 	u0slab = Expression(('u0*exp(-(pow((x[0] - xcenter),2)/(pow(sigma,2))))', '0'), u0=self.u0_SSE,
			# 	                    xcenter=self.xcenter, sigma=self.sigma_b, degree=1)
			# 	u0sub = Expression(('u0_sub*stick', '0'), u0_sub=self.u0_sub, stick=self.stick, degree=1)
			# 	# u0sub = Expression(('0', '0'), degree=1)
			# 	a_Boundary0 = assemble((1 / self.delta) * inner(self.T * (u0slab + u0sub), self.T * self.v) * self.ds(
			# 		self.slab))  # Prescribed slip boundary term
			#
			# elif self.event == 'sub':
			# 	u0slab = Expression(('u0_sub*stick', '0'), u0_sub=self.u0_sub, stick=self.stick, degree=1)
			# 	a_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) * self.ds(
			# 		self.slab))
			#
			# if self.event == 'sub_EQ' and i < self.sub_cycle_years:
			# 	u0slab = Expression(('u0_sub*stick', '0'), u0_sub=self.u0_sub, stick=self.stick, degree=1)
			# 	a_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) * self.ds(
			# 		self.slab))
			#
			# if self.event == 'sub_EQ' and i >= self.sub_cycle_years:  # EQ event
			# 	u0slab = Expression(('u0_sub*stick + u0d*exp(-(pow((x[0] - xcenter),2)/(pow(sigma,2))))', '0'),
			# 	                    u0d=self.u0_EQ, xcenter=self.xcenter, sigma=self.sigma_b,
			# 	                    u0_sub=self.u0_sub, stick=self.stick, degree=1)
			# 	a_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) * self.ds(
			# 		self.slab))

			self.b = Vector()
			self.a_Mass.init_vector(self.b, 0)
			self.L.mult(self.w0.vector(), self.b)
			if self.event != 'topo':
				self.b += a_Boundary0

			#self.kappa_update(dt)
			#self.point_source(dt)

			[bc.apply(self.b) for bc in self.bcs]
			[bc.apply(self.A) for bc in self.bcs]
			print("BC's applied", (time.time() - self.start) / 60.0, "minutes")

			self.w = Function(self.ME)
			self.solver.set_operator(self.A)
			self.solver.solve(self.w.vector(), self.b)
			print("System solved", (time.time() - self.start) / 60.0, "minutes")

			p, u = self.w.split()

			#self.k_frack(p)

			self.domain_flux = (-self.kappa / self.muf) * grad(p)  # element-wise flux vector
			self.bound_flux = assemble(inner(self.domain_flux, self.n) * self.ds(1))
			print("outward flux through seafloor: ", self.bound_flux)
			self.bound_flux2 = assemble(inner(self.domain_flux, self.n) * self.ds(2))
			print("outward flux through land surface: ", self.bound_flux2)

			self.q_velocity = Function(self.Q_gspace)
			self.q_velocity.assign(project(self.domain_flux, self.Q_gspace, solver_type='mumps'))
			print("Computed velocity field", (time.time() - self.start) / 60.0, "minutes")

			#self.v_strain = Function(self.Qspace)
			#self.v_strain.assign(project(nabla_div(u), self.Qspace, solver_type='mumps'))
			#self.psi = self.streamfunction()  # compute numerical streamfunction

			#esig_v = Expression('-rho_rock * g * x[1] - p', rho_rock=self.rho_rock, g=self.g, p=p, degree=1)
			#esigv = Function(self.Qspace)
			#esigv.assign(project(esig_v, self.Qspace, solver_type='mumps'))

			def get_eig(self, hes):
				mesh = hes.function_space().mesh()
				[eigL, eigR] = np.linalg.eig(
					hes.vector().array().reshape([4, self.pdim]).transpose().reshape([self.pdim, 2, 2]))

				eig = Function(self.Q_gspace)

				eig.vector().set_local(eigL.reshape([self.pdim * 2, 1], order='F').astype(float).flatten())

				return eig

			#stress = Function(self.Wspace)
			#stress.assign(project(self.sigma(u), self.Wspace, solver_type='mumps'))
			#effstress = Function(self.Wspace, solver_type='mumps')
			#effstress.assign(project(self.sigma(u) - p * self.I, self.Wspace, solver_type='mumps'))
			#eigstress = get_eig(self, stress)
			#eigeffstress = get_eig(self, effstress)

			####################### SAVE SOLUTION ###########################
			p.rename('p', 'p')
			self.pfile << p
			# p.rename('pnorm', 'pnorm')
			# self.pnormfile << pnorm
			u.rename('u', 'u')
			self.ufile << u
			self.p_velfile << self.q_velocity
			#stress.rename('stress', 'stress')
			# self.sigfile << stress
			# self.effsigfile << effstress
			# effstress.rename('effstress', 'effstress')
			# self.eigsigfile << eigstress
			# self.eigeffsigfile << eigeffstress
			# self.frackingfile << self.frack

			self.sea_flux_total[count] = self.bound_flux
			#self.sea_flux[:, count] = self.q_velocity.vector()[self.ocean_dof_pgrad]
			#self.flow_velocity[:, count] = self.q_velocity.vector()
			#self.Psi[:, count] = self.psi.vector()
			#self.vol_strain[:, count] = self.v_strain.vector()
			self.Sol_surf[:, count] = self.w.vector()[self.surface_dofs]
			#self.Sol_all[:, count] = self.w.vector()
			#self.esigv_slab[:, count] = esigv.vector()[self.dof_slabQ[0:self.size_p]]

			####################### UPDATE SOLUTION ###########################
			self.w0 = self.w

			count += 1
			print("Timestep:", count, "dt:", dt)

	def save_solution(self):
		if not os.path.exists(self.results_path):
			os.makedirs(self.results_path)
		if self.event == 'SSE':
			np.save(self.results_path + "Sol_surf_2D_SSE.npy", self.Sol_surf)
			np.save(self.results_path + "Sol_all_2D_SSE.npy", self.Sol_all)
			np.save(self.results_path + "sea_flux_2D_SSE.npy", self.sea_flux)
			np.save(self.var_path + "dtt_comsum_SSE.npy", self.dtt_comsum)
			np.save(self.var_path + "esigv_slab_SSE.npy", self.esigv_slab)
		elif self.event == 'sub':
			np.save(self.results_path + "Sol_surf_2D_sub.npy", self.Sol_surf)
			np.save(self.results_path + "Sol_all_2D_sub.npy", self.Sol_all)
			np.save(self.results_path + "sea_flux_2D_sub.npy", self.sea_flux)
			np.save(self.results_path + "flow_velocity_2D_sub.npy", self.flow_velocity)
			np.save(self.results_path + "vol_strain_2D_sub.npy", self.vol_strain)
			np.save(self.results_path + "Streamfun_2D_sub.npy", self.Psi)
			np.save(self.results_path + "sea_flux_total_2D_sub.npy", self.sea_flux_total)
			np.save(self.var_path + "dtt_comsum_sub.npy", self.dtt_comsum)
			np.save(self.var_path + "sub_cycle_sub.npy", self.sub_cycle_years)
		elif self.event == 'sub_EQ':
			np.save(self.results_path + "Sol_surf_2D_sub_EQ.npy", self.Sol_surf)
			np.save(self.results_path + "Sol_all_2D_sub_EQ.npy", self.Sol_all)
			np.save(self.results_path + "sea_flux_2D_sub_EQ.npy", self.sea_flux)
			np.save(self.var_path + "dtt_comsum_sub_EQ.npy", self.dtt_comsum)
			np.save(self.var_path + "sub_cycle_sub_EQ.npy", self.sub_cycle_years)
		elif self.event == 'EQ':
			if self.loop == 'yes':
				#np.save(self.results_path + "Sol_all_2D_EQ_loop_%s.npy" % self.loop_name, self.Sol_all)
				np.save(self.results_path + "Sol_surf_2D_EQ_loop_%s.npy" % self.loop_name, self.Sol_surf)
				np.save(self.results_path + "sea_flux_total_2D_EQ_loop_%s.npy" % self.loop_name, self.sea_flux_total)
				#np.save(self.results_path + "sea_flux_2D_EQ_loop_%s.npy" % self.loop_name, self.sea_flux)
				#np.save(self.results_path + "vol_strain_2D_EQ_loop_%s.npy" % self.loop_name, self.vol_strain)
				#np.save(self.results_path + "flow_velocity_2D_EQ_loop_%s.npy" % self.loop_name, self.flow_velocity)
				#np.save(self.results_path + "Streamfun_2D_EQ_loop_%s.npy" % self.loop_name, self.Psi)
				np.save(self.var_path + "dtt_comsum_EQ.npy", self.dtt_comsum)
			if self.loop == 'no':
				np.save(self.results_path + "Sol_surf_2D_EQ.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_all_2D_EQ.npy", self.Sol_all)
				np.save(self.results_path + "sea_flux_2D_EQ.npy", self.sea_flux)
				np.save(self.var_path + "dtt_comsum_EQ.npy", self.dtt_comsum)
				np.save(self.results_path + "flow_velocity_2D_EQ.npy", self.flow_velocity)
				np.save(self.results_path + "vol_strain_2D_EQ.npy", self.vol_strain)
				np.save(self.results_path + "Streamfun_2D_EQ.npy", self.Psi)
				np.save(self.results_path + "sea_flux_total_2D_EQ.npy", self.sea_flux_total)
		elif self.event == 'topo':
			if self.loop == 'yes':
				np.save(self.results_path + "Sol_surf_2D_topo_loop_%s.npy" % self.loop_name, self.Sol_surf)
				np.save(self.results_path + "Streamfun_2D_topo_loop_%s.npy" % self.loop_name, self.Psi)
				np.save(self.results_path + "sea_flux_total_2D_topo_loop_%s.npy" % self.loop_name,
				        self.sea_flux_total)
				np.save(self.results_path + "sea_flux_2D_topo_loop_%s.npy" % self.loop_name, self.sea_flux)
				np.save(self.results_path + "flow_velocity_2D_topo_loop_%s.npy" % self.loop_name,
				        self.flow_velocity)
				np.save(self.results_path + "vol_strain_2D_topo_loop_%s.npy" % self.loop_name, self.vol_strain)
				np.save(self.results_path + "Sol_all_2D_topo_loop_%s.npy" % self.loop_name, self.Sol_all)
				np.save(self.var_path + "dtt_comsum_topo.npy", self.dtt_comsum)
			if self.loop == 'no':
				np.save(self.results_path + "Sol_surf_2D_topo.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_all_2D_topo.npy", self.Sol_all)
				np.save(self.results_path + "sea_flux_2D_topo.npy", self.sea_flux)
				np.save(self.var_path + "dtt_comsum_topo.npy", self.dtt_comsum)
				np.save(self.results_path + "flow_velocity_2D_topo.npy", self.flow_velocity)
				np.save(self.results_path + "vol_strain_2D_topo.npy", self.vol_strain)
				np.save(self.results_path + "Streamfun_2D_topo.npy", self.Psi)
				np.save(self.results_path + "sea_flux_total_2D_topo.npy", self.sea_flux_total)

class Poroelasticity3DSynthetic:
	def __init__(self, data_file, origin, theta, ndim, mesh, plot_figs, event, new_mesh, permeability,
	             SSE_days,
	             sub_cycle_years, sigma_bx, xcenter, sigma_by, ycenter, SS_migratedip, SS_migratestrike, u0_EQdip,
	             u0_EQstrike, u0_SSEdip, u0_SSEstrike, u0_subdip, u0_substrike, surface_k, ocean_k, loop):


		print("yup, working")
		self.start = time.time()
		self.data_file = data_file
		self.origin = origin
		self.theta = theta
		self.mesh = mesh
		self.plot_figs = plot_figs
		self.event = event
		self.new_mesh = new_mesh
		self.permeability = permeability
		self.SSE_days = SSE_days
		self.sub_cycle_years = sub_cycle_years
		self.sigma_bx = sigma_bx
		self.xcenter = xcenter
		self.sigma_by = sigma_by
		self.ycenter = ycenter
		self.SS_migratex = SS_migratedip
		self.SS_migratey = SS_migratestrike
		self.u0_EQx = u0_EQdip
		self.u0_EQy = u0_EQstrike
		self.u0d_SSE = u0_SSEdip
		self.u0s_SSE = u0_SSEstrike
		self.u0_subx = u0_subdip
		self.u0_suby = u0_substrike
		self.surface_k = surface_k
		self.ocean_k = ocean_k
		# solver = PETScKrylovSolver("gmres","petsc_amg")
		# solver = PETScKrylovSolver("cg","ilu")
		#self.solver = LinearSolver('mumps')
		#self.prm = self.solver.parameters
		#self.prm['reuse_factorization'] = True  # Saves Factorization of LHS

		self.V = VectorElement('Lagrange', self.mesh.ufl_cell(), 2)
		self.Q = FiniteElement('Lagrange', self.mesh.ufl_cell(), 1)
		self.ME = FunctionSpace(self.mesh, MixedElement([self.Q, self.V]))
		self.Vspace = FunctionSpace(self.mesh, self.V)
		self.Qspace = FunctionSpace(self.mesh, self.Q)
		self.Q_gspace = VectorFunctionSpace(self.mesh, "Lagrange", 1)
		self.Wspace = TensorFunctionSpace(self.mesh, "Lagrange", 1)

		self.Kelement = FiniteElement('CG', self.mesh.ufl_cell(), 1)
		self.Kspace = FunctionSpace(self.mesh, self.Kelement)

		(self.p, self.u) = TrialFunctions(self.ME)
		(self.q, self.v) = TestFunctions(self.ME)
		self.ds = Measure("ds", domain=self.mesh)
		self.dx = Measure("dx", domain=self.mesh)#
		self.n = FacetNormal(self.mesh)  # normal vector
		self.dim = self.ME.dim()
		self.pdim = self.ME.sub(0).dim()
		self.udim = self.ME.sub(1).dim()
		self.geom_dim = ndim + 1
		
	

		##################### GPS STATION LOCATIONS ##############################
		wb = pyxl.load_workbook(self.data_file)
		gps_data = wb['GPS_data']
		
		self.x_gps, self.y_gps = latlon2XY(gps_data, self.origin, self.theta)
		self.z_gps = np.zeros(len(self.x_gps))

		self.results_path = 'results/numpy_results/'
		self.var_path = 'results/numpy_variables/'
		self.paraviewpath = 'results/paraview/'
		if not os.path.exists(self.paraviewpath):
			os.makedirs(self.paraviewpath)

		self.extract_coords()
		print("Coordinates extracted", (time.time() - self.start) / 60.0, "minutes")
		self.k_interpolation()
		self.plot_k()
		print("Permeability interpolated", (time.time() - self.start) / 60.0, "minutes")
		self.timestepping()
		
		######################### POROELASTIC PARAMETERS #############################
		self.delta = 1e-6  # Dirichlet control regularization on bottom boundary
		self.muf = Constant('1e-9')  # pore fluid viscosity [MPa s]
		self.B = .6  # Skempton's coefficient  ...Should this be constant?
		self.nuu = .38  # undrained poissons ratio
		self.nu = 0.27  # Poisson's ratio
		
		x = SpatialCoordinate(self.mesh)
		self.E = interpolate(347 * pow(-x[2], 0.516) + 5.324e4, self.Kspace)
		self.mu = interpolate(self.E / (2.0*(1.0 + self.nu)), self.Kspace)  # shear modulus
		self.lmbda = interpolate(self.E*self.nu / ((1.0 + self.nu)*(1.0 - 2.0*self.nu)), self.Kspace)  # Lame's parameter
		self.alpha = 3 * (self.nuu - self.nu) / (self.B * (1 - 2 * self.nu) * (1 + self.nuu))  # Biot's coefficient
		self.Se = interpolate((9*(self.nuu-self.nu)*(1-(2*self.nuu)))/
		                      (2*self.mu*pow(self.B,2)*(1-2*self.nu)*pow((1+self.nuu),2)), self.Kspace)  # mass specific storage
		
		
		self.kstick = 8e-5
		self.stick = Expression('1/(1+exp(-kstick*(1.1e5-x[0])))', kstick=self.kstick, degree=1)
		self.d = self.u.geometric_dimension()  # number of space dimensions
		self.I = Identity(self.d)
		self.T = (self.I - outer(self.n, self.n))  # tangent operator for boundary condition
		self.year = 60 * 60 * 24 * 365

		################### INITIAL AND BOUNDARY CONDITIONS ######################
		p0d = 0  # initial pressure
		self.w_init = Expression(('p0d', '0', '0', '0'), p0d=p0d, degree=1)
		self.w0 = Function(self.ME)
		self.w0.interpolate(self.w_init)
		self.ub0 = Expression(('0', '0', '0'), degree=1)
		self.ub = Expression(('1', '1', '1'), degree=1)
		self.pb0 = Expression('1', degree=1)
		self.bc1 = DirichletBC(self.ME.sub(1), self.ub0,  1)  # ocean surface
		self.bc2 = DirichletBC(self.ME.sub(1), self.ub,  2)  # land surface
		self.bc3 = DirichletBC(self.ME.sub(1), self.ub0,  3)  # subducting slab
		self.bc4 = DirichletBC(self.ME.sub(1), self.ub0,  4)  # ocean wedge
		self.bc5 = DirichletBC(self.ME.sub(1), self.ub0,  5)  # north side
		self.bc6 = DirichletBC(self.ME.sub(1), self.ub0,  6)  # back side (no slip)
		self.bc7 = DirichletBC(self.ME.sub(1), self.ub0,  7)  # south side
		self.bc8 = DirichletBC(self.ME.sub(1), self.ub0,  8)  # mantle

		self.bcp1 = DirichletBC(self.ME.sub(0), ('0'),  1)  # ocean surface (free flow condition)
		self.bcp2 = DirichletBC(self.ME.sub(0), ('1'),  2)  # land surface

		if self.event == 'topo':
			self.topo_flow_interp()
			self.bcp2 = DirichletBC(self.ME.sub(0), self.p_topo,  2)  # land surface
			self.bcs = [self.bc6, self.bc7, self.bcp1, self.bcp2]  # These are the BC's that are actually applied
		else:
			self.bcs = [self.bc6, self.bc7, self.bcp1]  # These are the BC's that are actually applied

		self.slab = 3  # the slab is labeled as boundary 3 for this mesh, but not always the case

		if self.event == 'SSE':
			self.u_addx = self.u0_SSEx
			self.u_addy = self.u0_SSEy
		elif self.event == 'sub':  # Amount of slip per day (at center of event)
			self.u_addx = self.u0_subx
			self.u_addy = self.u0_suby
		elif self.event == 'sub_EQ':  # Amount of slip per day (at center of event)
			self.u_addx = self.u0_subx
			self.u_addy = self.u0_suby

		################### SET UP TO SAVE SOLUTION  ######################


		self.Sol_all = np.empty((self.dim, self.nsteps))  # matrix to save entire solution
		self.Sol_surf = np.empty((self.surface_dofs.shape[0], self.nsteps))  # matrix to save surface solution
		self.Sol_gps = np.empty((self.x_gps.shape[0] * self.geom_dim, self.nsteps))  # matrix to save surface solution
		#elf.sea_flux = np.empty((self.ocean_dof_pgrad.shape[0], self.nsteps))  # ocean bottom solution
		#self.flow_velocity = np.empty((self.pdim * ndim, self.nsteps))  # ocean bottom solution
		#self.Psi = np.empty((self.size_p, self.nsteps))  # ocean bottom solution
		#self.vol_strain = np.empty((self.size_p, self.nsteps))  # ocean bottom solution
		self.sea_flux_total = np.empty(self.nsteps)  # matrix to save surface post-seismic solution
		print("Solution matrices created", (time.time() - self.start) / 60.0, "minutes")

		if loop == 'yes':
			if self.loop_variable == 'kappa':
				self.loop_name = "k" + str(-self.surface_k)
				pfilename = self.paraviewpath + "pressure2D_loop_k%s.pvd" % str(-self.surface_k)
				p_velfilename = self.paraviewpath + "pressure_vel2D_loop_k%s.pvd" % str(-self.surface_k)
				ufilename = self.paraviewpath + "def2D_loop_k%s.pvd" % str(-self.surface_k)
				sigfilename = self.paraviewpath + "stress2D_k%s.pvd" % str(-self.surface_k)
				effsigfilename = self.paraviewpath + "effstress2D_k%s.pvd" % str(-self.surface_k)
				eigsigfilename = self.paraviewpath + "eigstress2D_k%s.pvd" % str(-self.surface_k)
			elif self.loop_variable == 'sigma':
				self.loop_name = 'sig' + str(self.sigma_b)[:2]
				pfilename = self.paraviewpath + "pressure2D_loop_%s.pvd" % str(self.sigma_b)[:2]
				p_velfilename = self.paraviewpath + "pressure_vel2D_loop_%s.pvd" % str(self.sigma_b)[:2]
				ufilename = self.paraviewpath + "def2D_loop_%s.pvd" % str(self.sigma_b)[:2]
				sigfilename = self.paraviewpath + "stress2D_%s.pvd" % str(self.sigma_b)[:2]
				effsigfilename = self.paraviewpath + "effstress2D_%s.pvd" % str(self.sigma_b)[:2]
				eigsigfilename = self.paraviewpath + "eigstress2D_%s.pvd" % str(self.sigma_b)[:2]
			if self.loop_variable == 'B':
				self.loop_name = 'B' + str(-self.B)
				pfilename = self.paraviewpath + "pressure2D_loop_B%s.pvd" % str(-self.B)
				p_velfilename = self.paraviewpath + "pressure_vel2D_loop_B%s.pvd" % str(-self.B)
				ufilename = self.paraviewpath + "def2D_loop_k%s.pvd" % str(-self.B)
				sigfilename = self.paraviewpath + "stress2D_B%s.pvd" % str(-self.B)
				effsigfilename = self.paraviewpath + "effstress2D_B%s.pvd" % str(-self.B)
				eigsigfilename = self.paraviewpath + "eigstress2D_B%s.pvd" % str(-self.B)
		else:
			pfilename = self.paraviewpath + "pressure3D_syn.pvd"
			p_velfilename = self.paraviewpath + "pressure_vel3D_syn.pvd"
			ufilename = self.paraviewpath + "def3D_syn.pvd"
			sigfilename = self.paraviewpath + "stress3D_syn.pvd"
			effsigfilename = self.paraviewpath + "effstress3D_syn.pvd"
			eigsigfilename = self.paraviewpath + "eigstress3D_syn.pvd"
			eigeffsigfilename = self.paraviewpath + "eigeffstress3D_syn.pvd"

		self.pfile = File(pfilename)
		self.p_velfile = File(p_velfilename)
		self.ufile = File(ufilename)
		self.sigfile = File(sigfilename)
		self.effsigfile = File(effsigfilename)
		self.eigsigfile = File(eigsigfilename)
		# self.eigeffsigfile = File(eigeffsigfilename)


		################### SET UP, RUN, SOLVE and SAVE  ######################

		# self.plot_k()
		# self.plot_initial_cond()
		
		dt = self.dtt_comsum[0]
		self.assemble_system(dt)
		print("System assembled", (time.time() - self.start) / 60.0, "minutes")
		self.solve_system()

		#self.save_solution()

	def timestepping(self):
		if self.event == 'EQ':  # 15 min for 1 day, 4 hrs for 20 days, 2 days for 30 days, per month for 10 years
			# self.dtt = [15, 240, 288e1, 432e2]  # in minutes
			# self.ntime = [96, 120, 15, 122]
			self.dtt = [15]
			self.ntime = [3]
		elif self.event == 'SSE':  # 12 hrs for SSE days, 2 hr for 4 days, 12 hrs for total time-4-SSE_days
			self.dtt = [720, 120, 720]  # in minutes
			self.ntime = [self.SSE_days * 2, 48, (self.days - self.SSE_days - 4) * 2]
			self.dtt = [1]
			self.ntime = [2]
		elif self.event == 'sub_EQ':  # 1/year for 50 years followed by 'EQ' scheme
			self.dtt = [60 * 24 * 365, 15, 240, 288e1, 432e2]  # in minutes
			self.ntime = [self.sub_cycle_years, 96, 120, 15, 122]
			self.dtt = [1]
			self.ntime = [2]
		elif self.event == 'sub':  # 1/year for 50 years
			self.dtt = [60 * 24 * 365]  # in minutes
			self.ntime = [60 * 24 * 365 * self.sub_cycle_years]
			self.dtt = [1]
			self.ntime = [2]
		elif self.event == 'topo':  # 1/year for 100 years
			self.dtt = [60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years]
		self.dtt = [i * 60 for i in self.dtt]  # in seconds
		self.dtt_repeated = np.repeat(self.dtt, self.ntime)
		self.dtt_comsum = np.cumsum(self.dtt_repeated)
		self.dt = self.dtt_comsum[0]
		self.nsteps = self.dtt_comsum.size

	def k_interpolation(self):
		log_kr = -18.0
		alpha_k = 0.9

		#self.coords_K = self.K.tabulate_dof_coordinates().reshape((self.Kspace.dim(), -1))
		bctest1 = DirichletBC(self.Kspace, (1.0), 1)  # ocean surface
		bctest2 = DirichletBC(self.Kspace, (1.0), 2)  # land surface
		bctest3 = DirichletBC(self.Kspace, (1.0), 4)  # ocean surface near trench
		ktest = Function(self.Kspace)
		bctest1.apply(ktest)
		bctest2.apply(ktest)
		bctest3.apply(ktest)
		self.dof_surfK = np.where(ktest.vector().array() == 1.0)[0]
		

		if self.permeability == 'mapped':
			
			#interpolate surface permeability onto all nodes
			kappa_surf = kappakriging(self.ocean_k, self.data_file, self.origin, self.theta,
			                          self.xvec, self.yvec, self.zvec, self.dof_surfK)
			z_surf = surfkriging(self.xvec, self.yvec, self.zvec, self.dof_surfK)
			
			oceandof = np.where(z_surf < -10.)[0]
			kappa_surf[oceandof] = self.ocean_k
			self.kappa = Function(self.Kspace)
			self.kappa_vec = pow(10, (log_kr + ((kappa_surf - log_kr) *
			                                    pow((1 - (1e-3 * (self.zvec-z_surf))), -alpha_k))))
			self.kappa.dat.data[:] = self.kappa_vec

		elif self.permeability == 'constant':
			#self.kappa = KExpression3DConstant(self.surface_k, self.ocean_k,
			#                                   log_kr, alpha_k, self.coords_K, self.dof_surfK, degree=1)
			kappa = Expression('pow(10, (log_kr + ((k_surf - log_kr) * pow((1 - (1e-3 * (x[2]))), -alpha_k))))',
			                     log_kr=log_kr, k_surf = self.surface_k, alpha_k=alpha_k)
			#kappa = Expression('pow(10, -10)')
			self.kappa = interpolate(kappa,self.Kspace)
			
	def plot_k(self):
		self.kappafile = File(self.paraviewpath + 'kappa_3D.pvd')
		kappaplot = Function(self.Qspace)
		kappaplot.dat.data[:] = np.log10(self.kappa_vec)
		#kappaplot.interpolate(Expression('log10(kappa)', kappa=self.kappa, degree=1))
		# print(kappaplot.vector().min(), kappaplot.vector().max())
		self.kappafile.write(kappaplot)

	def strain(self, w):  # strain = 1/2 (grad u + grad u^T)
		return sym(nabla_grad(w))

	def sigma(self, w):  # stress = 2 mu strain + lambda tr(strain) I
		mu = interpolate(self.mu, self.Kspace)
		lmbda = interpolate(self.lmbda, self.Kspace)
		return 2.0 * mu * self.strain(w) + lmbda * div(w) * self.I
	
	def assemble_system(self, dt, assembleRHS=True):
		
		if assembleRHS:
			
			self.a_Mass = self.Se * self.p * self.q * dx  # Mass matrix
			self.a_Div = self.alpha*nabla_div(self.u) * self.q * dx  # Divergence matrix
			self.a_Grad = -self.alpha * self.p * (nabla_div(self.v)) * dx  # Gradient
			self.a_E = inner(self.sigma(self.u), nabla_grad(self.v)) * dx # Elasticity
			self.a_Boundary = (1 / self.delta) * dot(self.u, self.n) * dot(self.v, self.n)\
			                          * self.ds(self.slab) + (1 / self.delta) * inner(self.T * self.u, self.T * self.v) * self.ds(self.slab)
	
	
		self.a_K = dt * (self.kappa / self.muf) * inner(nabla_grad(self.p),
		                                                         nabla_grad(self.q)) * dx  # Stiffness matrix
		
		print("all LHS matrices assembled", (time.time() - self.start) / 60.0, "minutes")
		
		self.A = assemble(self.a_Mass + self.a_Div + self.a_K +
		                  self.a_Grad + self.a_E + self.a_Boundary, mat_type='aij')
		
		print("LHS assembled", (time.time() - self.start) / 60.0, "minutes")
		
		if assembleRHS:
			self.L = assemble(self.a_Mass + self.a_Div)
			print("RHS assembled", (time.time() - self.start) / 60.0, "minutes")

	def extract_coords(self):

		if not os.path.exists(self.var_path):
			os.makedirs(self.var_path)
		if self.new_mesh == 'yes':  # create surface/GPS indices and save them

			self.x_all, self.y_all, self.z_all =  SpatialCoordinate(self.mesh)
	
			xf = Function(self.Qspace)
			yf = Function(self.Qspace)
			zf = Function(self.Qspace)
			xm, ym, zm = SpatialCoordinate(self.mesh)
			
			self.xvec = xf.interpolate(xm).dat.data_ro
			self.yvec = yf.interpolate(ym).dat.data_ro
			self.zvec = zf.interpolate(zm).dat.data_ro

			bctest0 = DirichletBC(self.Qspace, (1), 1)  # ocean surface
			bctest1 = DirichletBC(self.Qspace, (2), 2)  # top surface

			ptest = Function(self.Qspace)
			bctest0.apply(ptest.vector())
			bctest1.apply(ptest.vector())
			self.ocean_dofs = np.where(ptest.vector().array() == 1)[0]
			self.surface_dofs = np.where(ptest.vector().array() == 2)[0]

			self.size_gps = self.x_gps.shape[0]
			size_w_gps = self.size_gps
			indices_gps = np.empty((self.size_gps, ))
			for j in range(0, self.size_gps):
				indice_gps = (np.abs(np.sqrt(pow((self.xvec[:] - self.x_gps[j]), 2) + pow((self.yvec[:] - self.y_gps[j]), 2)))).argmin()
				indices_gps[j] = indice_gps
				self.GPS_dofs = indices_gps

			np.save(self.var_path + "x_all_3D", self.xvec)
			np.save(self.var_path + "y_all_3D", self.yvec)
			np.save(self.var_path + "z_all_3D", self.zvec)
			np.save(self.var_path + "GPS_dofs_3D", self.GPS_dofs)
			np.save(self.var_path + "surface_dofs_3D", self.surface_dofs)
			np.save(self.var_path + "ocean_dofs_3D", self.ocean_dofs)
			#np.save(self.var_path + "size_p_3D", self.size_p)
			#np.save(self.var_path + "size_u_3D", self.size_u)

		elif self.new_mesh == 'no':  # load saved indices/variables
			self.x_all = np.load(self.var_path + "x_all_3D.npy")
			self.y_all = np.load(self.var_path + "y_all_3D.npy")
			self.z_all = np.load(self.var_path + "z_all_3D.npy")
			self.GPS_dofs = np.load(self.var_path + "GPS_dofs_3D.npy")
			self.surface_dofs = np.load(self.var_path + "surface_dofs_3D.npy")
			self.ocean_dofs = np.load(self.var_path + "ocean_dofs_3D.npy")
			self.size_p = np.load(self.var_path + "size_p_3D.npy")
			self.size_u = np.load(self.var_path + "size_u_3D.npy")

			self.ocean_dof_p = self.ocean_dofs[np.where(self.ocean_dofs < self.size_p)]
			self.ocean_dof_u = self.ocean_dofs[np.where(self.ocean_dofs >= self.size_p)] - self.size_p
			self.ocean_dof_pgrad = np.hstack((self.ocean_dof_p, (self.ocean_dof_p + self.ocean_dof_p.shape[0]),
			                                  (self.ocean_dof_p + 2 * self.ocean_dof_p.shape[0])))
			self.surface_dof_p = self.surface_dofs[np.where(self.surface_dofs < self.size_p)]
			self.surface_dof_u = self.surface_dofs[np.where(self.surface_dofs >= self.size_p)] - self.size_p
			self.surface_dof_pgrad = np.hstack((self.surface_dof_p, (self.surface_dof_p + self.surface_dof_p.shape[0]),
			                                    (self.surface_dof_p + 2 * self.surface_dof_p.shape[0])))

	def plot_initial_cond(self):  # plot the initial conditions
		plot(self.w0.sub(1), mode="glyphs", key="def")  # , rescale = False)
		plot(self.w0.sub(0), mode='color', key="pressure")  # , rescale = False)

	def streamfunction3D(self, constrained_domain=None):
		"""Stream function for a given 3D velocity field.
		The boundary conditions are weakly imposed through the term

			inner(q, grad(psi)*n)*ds,

		where u = curl(psi) is used to fill in for grad(psi) and set
		boundary conditions. This should work for walls, inlets, outlets, etc.
		"""

		vel = self.q_velocity
		qtest = TestFunction(self.Qspace)
		psi = TrialFunction(self.Qspace)

		a = inner(grad(qtest), grad(psi)) * dx - inner(qtest, dot(self.n, grad(psi))) * self.ds
		L = inner(grad(qtest), curl(vel)) * dx - dot(grad(qtest), cross(self.n, vel)) * self.ds
		# L = qtest* cross(self.n, vel) * self.ds

		psi_ = Function(self.Qspace)

		# Compute solution
		# A0 = assemble(inner(grad(qtest), grad(psi)) * dx)
		# L0 = inner(qtest, curl(vel)) * dx - dot(qtest, cross(self.n, vel)) * self.ds + inner(qtest, dot(self.n, grad(psi_))) * self.ds
		A = assemble(a)
		b = assemble(L)

		solver = LUSolver('petsc')

		b = assemble(L, tensor=b)
		solver.set_operator(A)
		# solver.set_nullspace(null_space)
		# null_space.orthogonalize(b)
		solver.solve(psi_.vector(), b)

		return psi_

	def solve_system(self):
		count = 0
		if self.event == 'EQ':  # For an EQ, this can be outside the time loop since we only use it for the first time step
			u0slab = Expression(('u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
			                    'u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
			                    '0'), u0d=self.u0_EQx, u0s=self.u0_EQy, xcenter=self.xcenter, ycenter=self.ycenter,
			                    sigmax=self.sigma_bx, sigmay=self.sigma_by, degree=1)
			# u0slab = Expression(('0','0','0'), degree = 1)
			u0slabf = Function(self.ME.sub(1))
			u0slabf.interpolate(u0slab)
			bndform = (1 / self.delta) * inner(u0slabf, self.T * self.v) * self.ds(
				self.slab)
			with assemble(bndform).dat.vec as vec:
				bnd_vec = vec
			
		for i in range(self.nsteps):
			
			if i == 0:
				dt = self.dtt_comsum[i]
				self.sol = Function(self.ME)
			else:
				dt = self.dtt_comsum[i] - self.dtt_comsum[i - 1]
			if i > 1:
				dt_old = self.dtt_comsum[i - 1] - self.dtt_comsum[i - 2]
				
			if i > 1 and dt != dt_old:  # Rebuild Left hand side when 'dt' changes
				print("re-assembling LHS", (time.time() - self.start) / 60.0, "minutes")
				self.assemble_system(dt, assembleRHS=False)
				self.prm['reuse_factorization'] = True  # Saves Factorization of LHS
			
			
			if i > 0:
				# Add slip to the boundary for a SSE and move center of slip along boundary
				if self.event == 'SSE' and count <= self.SSE_days:
					self.u0d_SSE += self.u_addx
					self.u0s_SSE += self.u_addy
					self.xcenter += self.SS_migratex
					self.ycenter += self.SS_migratey
				elif self.event == 'sub':
					self.u0_subx += self.u_addx
					self.u0_suby += self.u_addy
				elif self.event == 'sub_EQ':
					self.u0_subx += self.u_addx * (dt / self.year)
					self.u0_suby += self.u_addy * (dt / self.year)

			# Slip boundary Expression
			if self.event == 'SSE':
				u0slab = Expression((
					'u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
					'u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
					'0'), u0d=self.u0d_SSE, u0s=self.u0s_SSE, xcenter=self.xcenter,
					ycenter=self.ycenter, sigmax=self.sigma_bx, sigmay=self.sigma_by, degree=1)
				a_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) * self.ds(
					self.slab))
			elif self.event == 'sub':
				u0slab = Expression(('u0_subx*stick', 'u0_suby*stick', '0'), u0_subx=self.u0_subx,
				                    u0_suby=self.u0_suby, stick=self.stick, degree=1)
				a_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) * self.ds(
					self.slab))
			elif self.event == 'sub_EQ' and i < self.sub_cycle_years:
				u0slab = Expression(('u0_subx*stick', 'u0_suby*stick', '0'), u0_subx=self.u0_subx,
				                    u0_suby=self.u0_suby, stick=self.stick, degree=1)
				a_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) * self.ds(
					self.slab))
			elif self.event == 'sub_EQ' and i >= self.sub_cycle_years:
				u0slab = Expression(('u0_subx*stick + u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + '
				                     '(pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
				                     'u0_suby*stick + u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) '
				                     '+ (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
				                     '0'), u0d=self.u0_subx, u0s=self.u0_suby, xcenter=self.xcenter,
				                    ycenter=self.ycenter, sigmax=self.sigma_bx, sigmay=self.Tsigma_by, degree=1)
				a_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) * self.ds(
					self.slab))
			

			# here we move the RHS and slip BC to PETSc to multiply in solution from
			# previous timestep and add in slip boundary term
			RHS_mat = self.L.M.handle
			b = RHS_mat.getVecs()[0]
			with assemble(self.w0).dat.vec as vec:
				w0vec = vec
			RHS_mat.mult(w0vec, b)

			# transfer back to Function through a numpy array
			b += bnd_vec
			b_func = Function(self.ME)
			b_func.vector().set_local(b.array)
			
			print("Boundary condition set", (time.time() - self.start) / 60.0, "minutes")
			
			L = b_func
			A = self.A
			w = Function(self.ME)
			#Eps = MumpsSolver(self.E_state)
			
			params = {
                            "mat_solver_package": "mumps",
			          }
			
			# parameters = {"mat_type": "matfree",
			#               "snes_monitor": True,
			#               "ksp_type": "fgmres",
			#               "ksp_gmres_modifiedgramschmidt": True,
			#               "ksp_monitor_true_residual": True,
			#               "pc_type": "fieldsplit",
			#               "pc_fieldsplit_type": "schur",
			#               "pc_fieldsplit_schur_fact_type": "lower",
			#               "fieldsplit_0_ksp_type": "preonly",
			#               "fieldsplit_0_pc_type": "python",
			#               "fieldsplit_0_pc_python_type": "firedrake.AssembledPC",
			#               "fieldsplit_0_assembled_pc_type": "lu",
			#               "fieldsplit_1_ksp_type": "gmres",
			#               "fieldsplit_1_ksp_rtol": 1e-4,
			#               "fieldsplit_1_pc_type": "python",
			#               "fieldsplit_1_pc_python_type": "firedrake.PCDPC",
			#               "fieldsplit_1_pcd_Mp_ksp_type": "preonly",
			#               "fieldsplit_1_pcd_Mp_pc_type": "lu",
			#               "fieldsplit_1_pcd_Kp_ksp_type": "preonly",
			#               "fieldsplit_1_pcd_Kp_pc_type": "lu",
			#               "fieldsplit_1_pcd_Fp_mat_type": "matfree"}
			
			#aP = None #aP(self.ME) #preconditioner
			#P = None # assemble(aP, mat_type='nest')
			#solver = LinearSolver(A, P=P, solver_parameters=parameters)
			#solver.solve(w, b_func)
			
			solve(A, w, b_func, bcs=self.bcs, solver_parameters=params)
			#solve(a == L, self.w, bcs=self.bcs, solver_parameters={'ksp_type': 'cg'})
			print("System solved", (time.time() - self.start) / 60.0, "minutes")
			p, u = w.split()
			self.domain_flux = (-self.kappa / self.muf) * grad(p)  # element-wise flux vector
			self.bound_flux = assemble(inner(self.domain_flux, self.n) * self.ds(1))
			self.bound_flux2 = assemble(inner(self.domain_flux, self.n) * self.ds(2))
			print("outward flux through seafloor: ", self.bound_flux)
			print("outward flux through land surface: ", self.bound_flux2)
			#
			self.q_velocity = Function(self.Q_gspace)
			self.q_velocity.assign(project(self.domain_flux, self.Q_gspace))
			# print("Computed velocity field", (time.time() - self.start) / 60.0, "minutes")
			#
			# self.v_strain = Function(self.Qspace)
			# self.v_strain.assign(project(nabla_div(u), self.Qspace))
			# self.psi = self.streamfunction3D()  # compute numerical streamfunction
			#
			# self.stress = Function(self.Wspace)
			# self.stress.assign(project(self.sigma(u), self.Wspace))

			####################### SAVE SOLUTION ###########################
			p.rename('p', 'p')
			u.rename('u', 'u')
			q_velocity.rename('q', 'q')
			self.pfile.write(p)
			self.ufile.write(u)
			self.qfile.write(q)
			#self.ufile << u
			# self.sigfile << self.stress

			# self.sea_flux_total[count] = self.bound_flux
			# self.sea_flux[:, count] = self.q_velocity.vector()[self.ocean_dof_pgrad]
			# self.flow_velocity[:, count] = self.q_velocity.vector()
			# self.Psi[:, count] = self.psi.vector()
			# self.vol_strain[:, count] = self.v_strain.vector()
			# self.Sol_surf[:, count] = self.w.vector()[self.surface_dofs]
			# self.Sol_all[:, count] = self.w.vector()

			######################## PLOT SOLUTION ##########################
			if self.plot_figs == 'yes':
				HTML(X3DOM().html(p))
			# contour = plot(u)
			# contour = plot(p)


			####################### UPDATE SOLUTION ###########################
			self.w0 = w
			count += 1
			print("Timestep:", count)
			# plt.colorbar(contour)


			# time.sleep(1) # if you want the plots to pause for (# of sec) in between time steps, turn this on

	def save_solution(self):

		if not os.path.exists(self.results_path):
			os.makedirs(self.results_path)
		if self.permeability == 'mapped':
			if self.event == 'SSE':
				np.save(self.results_path + "Sol_surf_3D_SSE_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_SSE_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_SSE_mapped.npy", self.Sol_all)
			elif self.event == 'sub':
				np.save(self.results_path + "Sol_surf_3D_sub_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_mapped.npy", self.Sol_all)
			elif self.event == 'sub_EQ':
				np.save(self.results_path + "Sol_surf_3D_sub_EQ_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_EQ_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_EQ_mapped.npy", self.Sol_all)
			elif self.event == 'EQ':
				np.save(self.results_path + "Sol_surf_3D_EQ_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_EQ_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_EQ_mapped.npy", self.Sol_all)
				np.save(self.results_path + "vol_strain_3D_EQ_mapped.npy", self.vol_strain)
				np.save(self.results_path + "flow_velocity_3D_EQ_mapped.npy", self.flow_velocity)
				np.save(self.results_path + "sea_flux_3D_EQ_mapped.npy", self.sea_flux)
				np.save(self.results_path + "sea_flux_total_3D_EQ_mapped.npy", self.sea_flux_total)
		elif self.permeability == 'constant':
			if self.event == 'SSE':
				np.save(self.results_path + "Sol_surf_3D_SSE_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_SSE_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_SSE_constant.npy", self.Sol_all)
			elif self.event == 'sub':
				np.save(self.results_path + "Sol_surf_3D_sub_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_constant.npy", self.Sol_all)
			elif self.event == 'sub_EQ':
				np.save(self.results_path + "Sol_surf_3D_sub_EQ_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_EQ_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_EQ_constant.npy", self.Sol_all)
				np.save(self.results_path + "vol_strain_3D_sub_EQ_constant.npy", self.vol_strain)
				np.save(self.results_path + "flow_velocity_3D_sub_EQ_constant.npy", self.flow_velocity)
				np.save(self.results_path + "sea_flux_3D_sub_EQ_constant.npy", self.sea_flux)
				np.save(self.results_path + "sea_flux_total_3D_sub_EQ_constant.npy", self.sea_flux_total)
			elif self.event == 'EQ':
				np.save(self.results_path + "Sol_surf_3D_EQ_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_EQ_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_EQ_constant.npy", self.Sol_all)
				np.save(self.results_path + "vol_strain_3D_EQ_constant.npy", self.vol_strain)
				np.save(self.results_path + "flow_velocity_3D_EQ_constant.npy", self.flow_velocity)
				np.save(self.results_path + "sea_flux_3D_EQ_constant.npy", self.sea_flux)
				np.save(self.results_path + "sea_flux_total_3D_EQ_constant.npy", self.sea_flux_total)

class Poroelasticity3D:
	def __init__(self, data_file, origin, theta, ndim, mesh, plot_figs, EQtype, event, new_mesh, permeability,
	             sub_cycle_years, u0_subdip, u0_substrike, surface_k, ocean_k, loop, **kwargs):

		self.start = time.time()
		self.data_file = data_file
		self.origin = origin
		self.theta = theta
		self.mesh = mesh
		self.plot_figs = plot_figs
		self.EQtype = EQtype
		self.event = event
		self.new_mesh = new_mesh
		self.permeability = permeability
		self.sub_cycle_years = sub_cycle_years
		self.u0_subx = u0_subdip
		self.u0_suby = u0_substrike
		self.surface_k = surface_k
		self.ocean_k = ocean_k
	
		self.V = VectorElement('Lagrange', self.mesh.ufl_cell(), 2)
		self.Q = FiniteElement('Lagrange', self.mesh.ufl_cell(), 1)
		self.ME = FunctionSpace(self.mesh, MixedElement([self.Q, self.V]))
		self.Vspace = FunctionSpace(self.mesh, self.V)
		self.Qspace = FunctionSpace(self.mesh, self.Q)
		self.Q_gspace = VectorFunctionSpace(self.mesh, "Lagrange", 1)
		self.Wspace = TensorFunctionSpace(self.mesh, "Lagrange", 1)
		self.Kelement = FiniteElement('CG', self.mesh.ufl_cell(), 1)
		self.Kspace = FunctionSpace(self.mesh, self.Kelement)
		
		(self.p, self.u) = TrialFunctions(self.ME)
		(self.q, self.v) = TestFunctions(self.ME)
		self.ds = Measure("ds", domain=self.mesh)
		self.dx = Measure("dx", domain=self.mesh)#
		self.n = FacetNormal(self.mesh)  # normal vector
		self.dim = self.ME.dim()
		self.pdim = self.ME.sub(0).dim()
		self.udim = self.ME.sub(1).dim()
		self.geom_dim = ndim + 1


		##################### GPS STATION LOCATIONS ##############################
		wb = pyxl.load_workbook(self.data_file)
		gps_data = wb['GPS_data']
		
		self.x_gps, self.y_gps = latlon2XY(gps_data, self.origin, self.theta)
		self.z_gps = np.zeros(len(self.x_gps))
		
		self.results_path = 'results/numpy_results/'
		self.var_path = 'results/numpy_variables/'
		self.paraviewpath = 'results/paraview/'
		self.inverse_results = '/fenics/shared/inverse_codes/results_deterministic/numpy_results/'
		self.inverse_var = '/fenics/shared/inverse_codes/results_deterministic/numpy_variables/'
	
		if not os.path.exists(self.results_path):
			os.makedirs(self.results_path)
		if not os.path.exists(self.var_path):
			os.makedirs(self.var_path)
		if not os.path.exists(self.paraviewpath):
			os.makedirs(self.paraviewpath)

		self.extract_coords()
		print("Coordinates extracted", (time.time() - self.start) / 60.0, "minutes")
		self.k_interpolation()
		self.plot_k()
		print("Permeability interpolated", (time.time() - self.start) / 60.0, "minutes")
		self.timestepping()

		######################### POROELASTIC PARAMETERS #############################
		self.delta = 1e-6  # Dirichlet control regularization on bottom boundary
		self.muf = Constant('1e-9')  # pore fluid viscosity [MPa s]
		self.B = .6  # Skempton's coefficient  ...Should this be constant?
		self.nuu = .35  # undrained poissons ratio
		self.nu = 0.25  # Poisson's ratio
		x = SpatialCoordinate(self.mesh)
		self.E = interpolate(500 * pow(-x[2], 0.516) + 2e4, self.Kspace)
		#self.E = interpolate(347 * pow(-x[2], 0.516) + 5.324e4, self.Kspace)
		self.mu = interpolate(self.E / (2.0*(1.0 + self.nu)), self.Kspace)  # shear modulus
		self.lmbda = interpolate(self.E*self.nu / ((1.0 + self.nu)*(1.0 - 2.0*self.nu)), self.Kspace)  # Lame's parameter
		self.alpha = 3 * (self.nuu - self.nu) / (self.B * (1 - 2 * self.nu) * (1 + self.nuu))  # Biot's coefficient
		self.Se = interpolate((9*(self.nuu-self.nu)*(1-(2*self.nuu)))/
		                      (2*self.mu*pow(self.B,2)*(1-2*self.nu)*pow((1+self.nuu),2)), self.Kspace)  # mass specific storage

		self.kstick = 8e-5
		self.stick = Expression('1/(1+exp(-kstick*(1.1e5-x[0])))', kstick=self.kstick, degree=1)
		self.d = self.u.geometric_dimension()  # number of space dimensions
		self.I = Identity(self.d)
		self.T = (self.I - outer(self.n, self.n))  # tangent operator for boundary condition
		self.year = 60 * 60 * 24 * 365

		# Seplot = Function(self.Qspace)
		# Seplot.interpolate(Expression('S', S=self.Se, degree=1))
		# print(Seplot.vector().min(), Seplot.vector().max())



		################### INITIAL AND BOUNDARY CONDITIONS ######################
		####### Initial conditions: (('p0','ux0','uy0')) ##########
		#p0d = 0  # initial pressure
		#self.w_init = Expression(('p0d', '0', '0', '0'), p0d=p0d, degree=1)
		self.sol_old = Function(self.ME)
		#self.w0.interpolate(self.w_init)
		self.ub0 = Expression(('0', '0', '0'), degree=1)
		self.ub = Expression(('1', '1', '1'), degree=1)
		self.pb0 = Expression('1', degree=1)
		self.bc1 = DirichletBC(self.ME.sub(1), self.ub0,  1)  # ocean surface
		self.bc2 = DirichletBC(self.ME.sub(1), self.ub,  2)  # land surface
		self.bc3 = DirichletBC(self.ME.sub(1), self.ub0,  3)  # subducting slab
		#self.bc4 = DirichletBC(self.ME.sub(1), self.ub0,  4)  # ocean wedge
		self.bc5 = DirichletBC(self.ME.sub(1), self.ub0,  5)  # north side
		self.bc6 = DirichletBC(self.ME.sub(1), self.ub0,  6)  # back side (no slip)
		self.bc7 = DirichletBC(self.ME.sub(1), self.ub0,  7)  # south side
		self.bc8 = DirichletBC(self.ME.sub(1), self.ub0,  8)  # mantle

		self.bcp1 = DirichletBC(self.ME.sub(0), ('0.'),  1, method="geometric")  # ocean surface (free flow condition)
		self.bcp2 = DirichletBC(self.ME.sub(0), ('1'),  2)  # land surface

		if self.event == 'topo':
			self.topo_flow_interp()
			self.bcp2 = DirichletBC(self.ME.sub(0), self.p_topo,  2)  # land surface
			self.bcs = [self.bc7, self.bcp1, self.bcp2]  # These are the BC's that are actually applied
		else:
			self.bcs = [self.bc6, self.bc7, self.bcp1]  # These are the BC's that are actually applied

		self.slab = 3  # the slab is labeled as boundary 3 for this mesh, but not always the case

		if self.event == 'SSE':
			self.u_addx = self.u0_SSEx
			self.u_addy = self.u0_SSEy
		elif self.event == 'sub':  # Amount of slip per day (at center of event)
			self.u_addx = self.u0_subx
			self.u_addy = self.u0_suby
		elif self.event == 'sub_EQ':  # Amount of slip per day (at center of event)
			self.u_addx = self.u0_subx
			self.u_addy = self.u0_suby

		################### SET UP TO SAVE SOLUTION  ######################
		#self.Sol_all = np.empty((self.dim, self.nsteps))  # matrix to save entire solution
		#self.Sol_surf = np.empty((self.surface_dofs.shape[0], self.nsteps))  # matrix to save surface solution
		#elf.sea_flux = np.empty((self.ocean_dof_pgrad.shape[0], self.nsteps))  # ocean bottom solution
		#self.flow_velocity = np.empty((self.pdim * ndim, self.nsteps))  # ocean bottom solution
		#self.Psi = np.empty((self.size_p, self.nsteps))  # ocean bottom solution
		#self.vol_strain = np.empty((self.size_p, self.nsteps))  # ocean bottom solution
		self.sea_flux_total = np.empty(self.nsteps)  # matrix to save surface post-seismic solution
		self.Sol_surf_p = np.empty((self.surface_dofs.shape[0], self.nsteps))  # matrix to save surface solution
		self.Sol_gps = np.empty((self.x_gps.shape[0], ndim, self.nsteps))  # matrix to save surface solution


		if loop == 'yes':
			if self.loop_variable == 'kappa':
				self.loop_name = "k" + str(-self.surface_k)
				pfilename = self.paraviewpath + "pressure2D_loop_k%s.pvd" % str(-self.surface_k)
				p_velfilename = self.paraviewpath + "pressure_vel2D_loop_k%s.pvd" % str(-self.surface_k)
				ufilename = self.paraviewpath + "def2D_loop_k%s.pvd" % str(-self.surface_k)
				sigfilename = self.paraviewpath + "stress2D_k%s.pvd" % str(-self.surface_k)
				effsigfilename = self.paraviewpath + "effstress2D_k%s.pvd" % str(-self.surface_k)
				eigsigfilename = self.paraviewpath + "eigstress2D_k%s.pvd" % str(-self.surface_k)
			elif self.loop_variable == 'sigma':
				self.loop_name = 'sig' + str(self.sigma_b)[:2]
				pfilename = self.paraviewpath + "pressure2D_loop_%s.pvd" % str(self.sigma_b)[:2]
				p_velfilename = self.paraviewpath + "pressure_vel2D_loop_%s.pvd" % str(self.sigma_b)[:2]
				ufilename = self.paraviewpath + "def2D_loop_%s.pvd" % str(self.sigma_b)[:2]
				sigfilename = self.paraviewpath + "stress2D_%s.pvd" % str(self.sigma_b)[:2]
				effsigfilename = self.paraviewpath + "effstress2D_%s.pvd" % str(self.sigma_b)[:2]
				eigsigfilename = self.paraviewpath + "eigstress2D_%s.pvd" % str(self.sigma_b)[:2]
			if self.loop_variable == 'B':
				self.loop_name = 'B' + str(-self.B)
				pfilename = self.paraviewpath + "pressure2D_loop_B%s.pvd" % str(-self.B)
				p_velfilename = self.paraviewpath + "pressure_vel2D_loop_B%s.pvd" % str(-self.B)
				ufilename = self.paraviewpath + "def2D_loop_k%s.pvd" % str(-self.B)
				sigfilename = self.paraviewpath + "stress2D_B%s.pvd" % str(-self.B)
				effsigfilename = self.paraviewpath + "effstress2D_B%s.pvd" % str(-self.B)
				eigsigfilename = self.paraviewpath + "eigstress2D_B%s.pvd" % str(-self.B)
		else:
			pfilename = self.paraviewpath + "pressure3D.pvd"
			p_velfilename = self.paraviewpath + "flux0_comp3D.pvd"
			p_velfilename1 = self.paraviewpath + "flux1_comp3D.pvd"
			ufilename = self.paraviewpath + "def3D.pvd"
			sigfilename = self.paraviewpath + "stress3D.pvd"
			effsigfilename = self.paraviewpath + "effstress3D.pvd"
			eigsigfilename = self.paraviewpath + "eigstress3D.pvd"
			eigeffsigfilename = self.paraviewpath + "eigeffstress3D.pvd"

		self.pfile = File(pfilename)
		self.qfile = File(p_velfilename)
		self.qfile1 = File(p_velfilename1)
		self.ufile = File(ufilename)
		self.sigfile = File(sigfilename)
		self.effsigfile = File(effsigfilename)
		self.eigsigfile = File(eigsigfilename)
		# self.eigeffsigfile = File(eigeffsigfilename)
		print("Solution matrices created", (time.time() - self.start) / 60.0, "minutes")

		################### SET UP, RUN, SOLVE and SAVE  ######################
		self.slip_interpolation()
		print("Slip data interpolated", (time.time() - self.start) / 60.0, "minutes")
		# self.plot_topo()
		self.plot_slip()
		# self.plot_initial_cond()
		dt = self.dtt_comsum[0]
		self.assemble_system(dt)

		print("System assembled", (time.time() - self.start) / 60.0, "minutes")
		self.solve_system()
		self.save_solution()
	
	def extract_coords(self):
		
		if not os.path.exists(self.var_path):
			os.makedirs(self.var_path)
		if self.new_mesh == 'yes':  # create surface/GPS indices and save them
			
			self.x_all, self.y_all, self.z_all = SpatialCoordinate(self.mesh)
			
			xf = Function(self.Qspace)
			yf = Function(self.Qspace)
			zf = Function(self.Qspace)
			xm, ym, zm = SpatialCoordinate(self.mesh)
			
			self.xvec = xf.interpolate(xm).dat.data_ro
			self.yvec = yf.interpolate(ym).dat.data_ro
			self.zvec = zf.interpolate(zm).dat.data_ro
			
			bctest0 = DirichletBC(self.Qspace, (1), 1)  # ocean surface
			bctest1 = DirichletBC(self.Qspace, (2), 2)  # top surface
			bctest2 = DirichletBC(self.Qspace, (3), 3)  # slab surface

			
			ptest = Function(self.Qspace)
			bctest0.apply(ptest.vector())
			bctest1.apply(ptest.vector())
			bctest2.apply(ptest.vector())
			self.ocean_dofs = np.where(ptest.vector().array() == 1)[0]
			self.surface_dofs = np.where(ptest.vector().array() == 2)[0]
			self.slab_dofs = np.where(ptest.vector().array() == 3)[0]


			self.size_gps = self.x_gps.shape[0]
			size_w_gps =  self.size_gps
			indices_gps = np.empty((self.size_gps, ))
			for j in range(0, self.size_gps):
				indice_gps = (np.abs(np.sqrt(pow((self.xvec[:] - self.x_gps[j]), 2)
				                             + pow((self.yvec[:] - self.y_gps[j]), 2)
				                             + pow((self.zvec[:] - 0.), 2)))).argmin()
				indices_gps[j] = indice_gps
			self.GPS_dofs = indices_gps.astype('int')
			
			
			np.save(self.var_path + "x_all_3D", self.xvec)
			np.save(self.var_path + "y_all_3D", self.yvec)
			np.save(self.var_path + "z_all_3D", self.zvec)
			np.save(self.var_path + "GPS_dofs_3D", self.GPS_dofs)
			np.save(self.var_path + "surface_dofs_3D", self.surface_dofs)
			np.save(self.var_path + "ocean_dofs_3D", self.ocean_dofs)
	
	def timestepping(self):
		if self.event == 'EQ':  # 15 min for 1/2 day, 4 hrs for 10 days, 12 hours for 30 days, 2 days for 2 months, per month for 10 years, per year for 50 years
			self.dtt = [15, 240, 720, 288e1, 2 * 432e2]  # , 5256e2]  # in minutes
			self.ntime = [48, 60, 60, 30, 60]  # , 50]
			# self.dtt = [15, 240, 720, 288e1, 432e2]#, 5256e2]  # in minutes
			# self.ntime = [3, 3, 3, 3, 3]#, 50]
			#self.dtt = [15, 240, 288e1, 432e2]#, 5256e2]  # in minutes
			#self.ntime = [2, 2, 2, 2]#, 50]
			#self.dtt = [15]
			#self.ntime = [1]
		elif self.event == 'SSE':  # 12 hrs for SSE days, 2 hr for 4 days, 12 hrs for total time-4-SSE_days
			self.dtt = [720, 120, 720]  # in minutes
			self.ntime = [self.SSE_days * 2, 48, (self.days - self.SSE_days - 4) * 2]
			self.dtt = [1]
			self.ntime = [2]
		elif self.event == 'sub_EQ':  # 1/year for 50 years followed by 'EQ' scheme
			self.dtt = [60 * 24 * 365, 15, 240, 288e1, 432e2]  # in minutes
			self.ntime = [self.sub_cycle_years, 96, 120, 15, 122]
			self.dtt = [1]
			self.ntime = [2]
		elif self.event == 'sub':  # 1/year for 50 years
			self.dtt = [60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years]
		# self.dtt = [1]
		# self.ntime = [2]
		elif self.event == 'topo':  # 1/year for 100 years
			self.dtt = [60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years]
		self.dtt = [i * 60 for i in self.dtt]  # in seconds
		self.dtt_repeated = np.repeat(self.dtt, self.ntime)
		self.dtt_comsum = np.cumsum(self.dtt_repeated)
		self.nsteps = self.dtt_comsum.size
	
	def k_interpolation(self):
		log_kr = -18.0
		alpha_k = 0.9
		
		# self.coords_K = self.K.tabulate_dof_coordinates().reshape((self.Kspace.dim(), -1))
		bctest1 = DirichletBC(self.Kspace, (1.0), 1)  # ocean surface
		bctest2 = DirichletBC(self.Kspace, (1.0), 2)  # land surface
		ktest = Function(self.Kspace)
		bctest1.apply(ktest)
		bctest2.apply(ktest)
		#bctest3.apply(ktest)
		self.dof_surfK = np.where(ktest.vector().array() == 1.0)[0]
		
		if self.permeability == 'mapped':
			
			# interpolate surface permeability onto all nodes
			kappa_surf = kappakriging(self.ocean_k, self.data_file, self.origin, self.theta,
			                          self.xvec, self.yvec, self.zvec, self.dof_surfK)
			z_surf = surfkriging(self.xvec, self.yvec, self.zvec, self.dof_surfK)
			
			oceandof = np.where(z_surf < -10.)[0]
			kappa_surf[oceandof] = self.ocean_k
			self.kappa = Function(self.Kspace)
			self.kappa_vec = pow(10, (log_kr + (((kappa_surf+1.) - log_kr) *
			                                    pow((1 - (1e-3 * 0.1*(self.zvec - z_surf))), -alpha_k))))
			self.kappa.dat.data[:] = self.kappa_vec
		
		elif self.permeability == 'constant':
			# self.kappa = KExpression3DConstant(self.surface_k, self.ocean_k,
			#                                   log_kr, alpha_k, self.coords_K, self.dof_surfK, degree=1)
			kappa = Expression('pow(10, (log_kr + ((k_surf - log_kr) * pow((1 - (1e-3 * (x[2]))), -alpha_k))))',
			                   log_kr=log_kr, k_surf=self.surface_k, alpha_k=alpha_k)
			# kappa = Expression('pow(10, -10)')
			self.kappa = interpolate(kappa, self.Kspace)
			
	def topo_flow_interp(self):

		self.coords_K = self.Kspace.tabulate_dof_coordinates().reshape((self.Kspace.dim(), -1))
		bctest1 = DirichletBC(self.Kspace, (1),  1)  # ocean surface
		bctest2 = DirichletBC(self.Kspace, (1),  2)  # land surface
		qtest = Function(self.Kspace)
		bctest1.apply(qtest.vector())
		bctest2.apply(qtest.vector())
		self.dof_surfK = np.where(qtest.vector() == 1)[0]

		wb = pyxl.load_workbook(self.data_file)
		topo_data = wb['topo_data3D']

		topo_p_surf = topokriging(topo_data, self.origin, self.theta,
		                          self.coords_K, self.dof_surfK, self.var_path)

		# oceanz = np.where(self.zsurfall_p < 0.0)[0]
		low_p = np.where(topo_p_surf < 0.02)[0]
		# self.psurfall[oceanz] = 0.0
		topo_p_surf[low_p] = 0.02

		# NEW method
		p_surf_func = Function(self.Kspace)
		p_surf_func.vector()[self.dof_surfK] = topo_p_surf
		self.p_topo = Expression(('ptopo'), ptopo=p_surf_func, degree=1)
	
	def slip_interpolation(self):
		tic = time.time()
		
		if self.EQtype == 'inversion':
			
			u0load = np.load(self.inverse_results + 'u0_slab_array.npy')
			size_u = int(u0load.shape[0]/3)
			u0load = u0load.reshape((size_u, 3), order = 'F')
			XYZslab = np.load(self.inverse_var + 'XYZ_slab.npy')[0:size_u, :]
			
			u0slab_interp = slabkriging(XYZslab, u0load, self.xvec, self.yvec, self.zvec, self.slab_dofs)

			self.u0slab = Function(self.Q_gspace)
			self.u0slab.dat.data[:] = u0slab_interp

		
		elif self.EQtype == 'data':
			
			# import from excel file ##
			wb = pyxl.load_workbook(self.data_file)
			slip_data = wb['slip_data']
			self.u0slab = SlipExpression(slip_data, self.origin, self.theta, self.coords_V, dof_slab, self.var_path,
			                             degree=1)
	
	def SSE_interpolation(self, timestep):

		bctest1 = DirichletBC(self.Vspace, (1, 1, 1),  3)  # slab surface
		stest = Function(self.Vspace)
		bctest1.apply(stest.vector())
		dof_slab = np.where(stest.vector() == 1)[0]

		self.u0_SSE = SSEExpression(self.coords_V, dof_slab, degree=1)

	def plot_k(self):
		self.kappafile = File(self.paraviewpath + 'kappa_3D.pvd')
		kappaplot = Function(self.Qspace)
		kappaplot.dat.data[:] = np.log10(self.kappa_vec)
		#kappaplot.interpolate(Expression('log10(kappa)', kappa=self.kappa, degree=1))
		# print(kappaplot.vector().min(), kappaplot.vector().max())
		self.kappafile.write(kappaplot)
		
	def plot_topo(self):
		self.topofile = File(self.paraviewpath + 'topo_3D.pvd')
		topoplot = Function(self.Qspace)
		topoplot.interpolate(Expression('p', p=self.p_topo, degree=1))
		# print(kappaplot.vector().min(), kappaplot.vector().max())
		self.topofile << topoplot
	
	def plot_slip(self):
		self.slipfile = File(self.paraviewpath + 'slip_3D.pvd')
		if self.EQtype == 'inversion':
			slipplot = self.u0slab
			# slipplot = self.ufwdV
			#slipplot = Function(self.Vspace)
			#slipplot.interpolate(self.u0slab)
		else:
			slipplot = Function(self.Vspace)
			slipplot.interpolate(self.u0slab)
		# print kappaplot.vector().min(), kappaplot.vector().max()
		self.slipfile.write(slipplot)
	
	def strain(self, w):  # strain = 1/2 (grad u + grad u^T)
		return sym(nabla_grad(w))

	def sigma(self, w):  # stress = 2 mu strain + lambda tr(strain) I
		return 2.0 * self.mu * self.strain(w) + self.lmbda * div(w) * self.I
	
	def assemble_system(self, dt, assembleRHS=True):
		mat_type = 'aij'
		if assembleRHS:
			self.a_Mass = self.Se * self.p * self.q * dx  # Mass matrix
			self.a_Div = self.alpha * nabla_div(self.u) * self.q * dx  # Divergence matrix
			self.a_Grad = -self.alpha * self.p * (nabla_div(self.v)) * dx  # Gradient
			self.a_E = inner(self.sigma(self.u), nabla_grad(self.v)) * dx  # Elasticity
			self.a_Boundary = (1 / self.delta) * dot(self.u, self.n) * dot(self.v, self.n) \
			                  * self.ds(self.slab) + (1 / self.delta) * inner(self.T * self.u,
			                                                                  self.T * self.v) * self.ds(self.slab)
		
		self.a_K = dt * (self.kappa / self.muf) * inner(nabla_grad(self.p),
		                                                nabla_grad(self.q)) * dx  # Stiffness matrix
		
		print("all LHS matrices assembled", (time.time() - self.start) / 60.0, "minutes")
		
		self.A = assemble(self.a_Mass + self.a_Div + self.a_K +
		                  self.a_Grad + self.a_E + self.a_Boundary, bcs=self.bcs, mat_type=mat_type)
		
		print("LHS assembled", (time.time() - self.start) / 60.0, "minutes")
		
		if assembleRHS:
			self.L = assemble(self.a_Mass + self.a_Div, bcs=self.bcs, mat_type=mat_type)
			print("RHS assembled", (time.time() - self.start) / 60.0, "minutes")
			
	def streamfunction3D(self, constrained_domain=None):
		"""Stream function for a given 3D velocity field.
		The boundary conditions are weakly imposed through the term

			inner(q, grad(psi)*n)*ds,

		where u = curl(psi) is used to fill in for grad(psi) and set
		boundary conditions. This should work for walls, inlets, outlets, etc.
		"""

		vel = self.q_velocity
		qtest = TestFunction(self.Qspace)
		psi = TrialFunction(self.Qspace)
		a = inner(grad(qtest), grad(psi)) * dx - inner(qtest, dot(self.n, grad(psi))) * self.ds
		L = inner(grad(qtest), curl(vel)) * dx - dot(grad(qtest), cross(self.n, vel)) * self.ds
		psi_ = Function(self.Qspace)
		A = assemble(a)
		b = assemble(L)
		solver = LUSolver('petsc')
		b = assemble(L, tensor=b)
		solver.set_operator(A)
		solver.solve(psi_.vector(), b)

		return psi_

	def test_PSD(self):

		def is_pd(x):
			return np.all(np.linalg.eigvals(x) > 0)

		def is_psd(x, tol=1e-5):
			E, V = np.linalg.eigh(x)
			return np.all(E > -tol)

		def is_symmetric(x):
			return (x.transpose() == x).all()

		A = self.A.array()

		print('A is {}'.format('symmetric' if is_symmetric(A) else ('not symmetric')))

		np.linalg.cholesky(A)

	def solve_system(self):
		count = 0
		if self.event == 'EQ':  # For an EQ, this can be outside the time loop since we only use it for the first time step
			bndform = (1 / self.delta) * inner(self.u0slab, self.T * self.v) * self.ds(
				self.slab)
			with assemble(bndform).dat.vec as vec:
				bnd_vec = vec
				
		# set up solvers and preconditioners
		params = {
			# 'ksp_view': True,
			#'ksp_type': 'gmres',
			'ksp_type': 'preonly',
			"ksp_monitor": True,
			'pc_type': 'lu',
			"pc_factor_mat_solver_package": "mumps",

			#'pc_factor_reuse_ordering': True,  # PCFactorSetReuseOrdering
			#'pc_factor_reuse_fill': True,

			# "ksp_monitor_true_residual": True,
			'ksp_initial_guess_nonzero': True,

			# "pc_type": "fieldsplit",
			# # p - DG0
			# "fieldsplit_0_ksp_type": "preonly",
			# "fieldsplit_0_pc_type": "lu",
			# "fieldsplit_0_pc_factor_mat_solver_package": "mumps",
			# "fieldsplit_0_pc_factor_reuse_factorization": True,
			#
			# # u - vec CG2			#
			# "fieldsplit_1_ksp_type": "preonly",
			# "fieldsplit_1_pc_type": "lu",
			# "fieldsplit_1_pc_factor_mat_solver_package": "mumps",
			#
			# # q - RT1
			# "fieldsplit_2_ksp_type": "preonly",
			# "fieldsplit_2_pc_type": "lu",
			# "fieldsplit_2_pc_factor_mat_solver_package": "mumps",

		}
		
		A = self.A
		solver = LinearSolver(A, solver_parameters=params)
		self.sol = Function(self.ME)
		dt = self.dtt_comsum[0]
		
		
		for i in range(self.nsteps):
			if i == 0:
				dt = self.dtt_comsum[i]
			else:
				dt = self.dtt_comsum[i] - self.dtt_comsum[i - 1]
			if i > 1:
				dt_old = self.dtt_comsum[i - 1] - self.dtt_comsum[i - 2]
			if i > 1 and dt != dt_old:  # Rebuild Left hand side when 'dt' changes
				print("re-assembling LHS", (time.time() - self.start) / 60.0, "minutes")
				self.assemble_system(dt, assembleRHS=False)
				A = self.A
				solver = LinearSolver(A, solver_parameters=params)
			if i > 0:
				# Add slip to the boundary for a SSE and move center of slip along boundary
				if self.event == 'SSE' and count <= self.SSE_days:
					print("This doesn't work yet...")
					quit()
					self.u0_SSE = SSE_interpolation(i)
				elif self.event == 'sub':
					self.u0_subx += self.u_addx * (dt / self.year)
					self.u0_suby += self.u_addy * (dt / self.year)
				elif self.event == 'sub_EQ':
					self.u0_subx += self.u_addx * (dt / self.year)
					self.u0_suby += self.u_addy * (dt / self.year)
			
			# Slip boundary Expression
			if self.event == 'SSE':
				u0slab = Expression((
					'u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
					'u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
					'0'), u0d=self.u0d_SSE, u0s=self.u0s_SSE, xcenter=self.xcenter,
					ycenter=self.ycenter, sigmax=self.sigma_bx, sigmay=self.sigma_by, degree=1)
				a_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) * self.ds(
					self.slab))
			elif self.event == 'sub':
				u0slab = Expression(('u0_subx*stick', 'u0_suby*stick', '0'), u0_subx=self.u0_subx,
				                    u0_suby=self.u0_suby, stick=self.stick, degree=1)
				a_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) * self.ds(
					self.slab))
			elif self.event == 'sub_EQ' and i < self.sub_cycle_years:
				u0slab = Expression(('u0_subx*stick', 'u0_suby*stick', '0'), u0_subx=self.u0_subx,
				                    u0_suby=self.u0_suby, stick=self.stick, degree=1)
				a_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) * self.ds(
					self.slab))
			elif self.event == 'sub_EQ' and i >= self.sub_cycle_years:
				u0slab = Expression(('u0_subx*stick + u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + '
				                     '(pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
				                     'u0_suby*stick + u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) '
				                     '+ (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
				                     '0'), u0d=self.u0_subx, u0s=self.u0_suby, xcenter=self.xcenter,
				                    ycenter=self.ycenter, sigmax=self.sigma_bx, sigmay=self.Tsigma_by, degree=1)
				a_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) * self.ds(
					self.slab))

			# here we move the RHS and slip BC to PETSc to multiply in solution from
			# previous timestep and add in slip boundary term
			RHS_mat = self.L.M.handle
			b = RHS_mat.getVecs()[0]
			with assemble(self.sol_old).dat.vec as vec:
				w0vec = vec
			RHS_mat.mult(w0vec, b)
			
			# transfer back to Function through a numpy array
			b += bnd_vec
			b_func = Function(self.ME)
			b_func.vector().set_local(b.array)
			[bc.apply(b_func) for bc in self.bcs]
			
			print("Boundary condition set", (time.time() - self.start) / 60.0, "minutes")
			
			solver.solve(self.sol, b_func)
			print("System solved", (time.time() - self.start) / 60.0, "minutes")
			p, u = self.sol.split()
			self.domain_flux = (-self.kappa / self.muf) * grad(p)  # element-wise flux vector
			self.bound_flux = assemble(inner(self.domain_flux, self.n) * self.ds(1))
			self.bound_flux2 = assemble(inner(self.domain_flux, self.n) * self.ds(2))
			print("outward flux through seafloor: ", self.bound_flux)
			print("outward flux through land surface: ", self.bound_flux2)
			#
			q = Function(self.Q_gspace)
			q.assign(project(self.domain_flux, self.Q_gspace))
			print("Computed velocity field", (time.time() - self.start) / 60.0, "minutes")
			
			u_V1 = Function(self.Q_gspace)
			u_V1.assign(project(u, self.Q_gspace))
			size_u = u_V1.vector().array().shape[0]
			
			# self.v_strain = Function(self.Qspace)
			# self.v_strain.assign(project(nabla_div(u), self.Qspace))
			# self.psi = self.streamfunction3D()  # compute numerical streamfunction
			#
			# self.stress = Function(self.Wspace)
			# self.stress.assign(project(self.sigma(u), self.Wspace))
			
			####################### SAVE SOLUTION ###########################
			p.rename('p', 'p')
			u.rename('u', 'u')
			q.rename('q', 'q')
			self.pfile.write(p)
			self.ufile.write(u)
			self.qfile.write(q)
			
			self.sea_flux_total[count] = self.bound_flux
			# self.sea_flux[:, count] = self.q_velocity.vector()[self.ocean_dof_pgrad]
			# self.flow_velocity[:, count] = self.q_velocity.vector()
			# self.Psi[:, count] = self.psi.vector()
			# self.vol_strain[:, count] = self.v_strain.vector()
			#self.Sol_surf[:, count] = self.w.vector()[self.surface_dofs]
			#self.Sol_all[:, count] = self.w.vector()
			#self.Sol_gps[:, count] = self.w.vector()[self.GPS_dofs]
			self.Sol_surf_p[:, count] = p.vector().array()[self.surface_dofs]
			#self.Sol_gps[:, count] = u_V1.vector().array()[self.GPS_dofs]
			self.Sol_gps[:, :, count] = u_V1.dat.data[self.GPS_dofs, :]#.reshape((size_u,), order = 'F')


			######################## PLOT SOLUTION ##########################
	
			# contour = plot(u)
			# contour = plot(p)


			####################### UPDATE SOLUTION ###########################
			self.sol_old = self.sol
			count += 1
			print("Timestep:", count)
			# plt.colorbar(contour)

	def save_solution(self):
		if not os.path.exists(self.results_path):
			os.makedirs(self.results_path)
		if self.permeability == 'mapped':

			if self.event == 'SSE':
				np.save(self.results_path + "Sol_surf_3D_SSE_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_SSE_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_SSE_mapped.npy", self.Sol_all)
			elif self.event == 'sub':
				np.save(self.results_path + "Sol_surf_3D_sub_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_mapped.npy", self.Sol_all)
			elif self.event == 'sub_EQ':
				np.save(self.results_path + "Sol_surf_3D_sub_EQ_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_EQ_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_EQ_mapped.npy", self.Sol_all)
			elif self.event == 'EQ':
				# np.save(self.var_path + "dtt_comsum_EQ.npy", self.dtt_comsum)
				np.save(self.results_path + "Sol_surfp_3D_EQ_mapped.npy", self.Sol_surf_p)
				np.save(self.results_path + "Sol_gps_3D_EQ_mapped.npy", self.Sol_gps)
				# np.save(self.results_path + "Sol_all_3D_EQ_mapped.npy", self.Sol_all)
				# np.save(self.results_path + "vol_strain_3D_EQ_mapped.npy", self.vol_strain)
				# np.save(self.results_path + "flow_velocity_3D_EQ_mapped.npy", self.flow_velocity)
				# np.save(self.results_path + "psi_3D_EQ_mapped.npy", self.Psi)
				# np.save(self.results_path + "sea_flux_3D_EQ_mapped.npy", self.sea_flux)
				np.save(self.results_path + "sea_flux_total_3D_EQ_mapped.npy", self.sea_flux_total)
			elif self.event == 'topo':
				np.save(self.results_path + "Sol_surf_3D_topo_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_all_3D_topo_mapped.npy", self.Sol_all)
				np.save(self.results_path + "sea_flux_3D_topo_mapped.npy", self.sea_flux)
				np.save(self.var_path + "dtt_comsum_topo_mapped.npy", self.dtt_comsum)
				np.save(self.results_path + "flow_velocity_3D_topo_mapped.npy", self.flow_velocity)
				np.save(self.results_path + "vol_strain_3D_topo_mapped.npy", self.vol_strain)
				np.save(self.results_path + "Streamfun_3D_topo_mapped.npy", self.Psi)
				np.save(self.results_path + "sea_flux_total_3D_topo_mapped.npy", self.sea_flux_total)
		elif self.permeability == 'constant':
			if self.event == 'SSE':
				np.save(self.results_path + "Sol_surf_3D_SSE_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_SSE_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_SSE_constant.npy", self.Sol_all)
			elif self.event == 'sub':
				np.save(self.results_path + "Sol_surf_3D_sub_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_constant.npy", self.Sol_all)
			elif self.event == 'sub_EQ':
				np.save(self.results_path + "Sol_surf_3D_sub_EQ_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_EQ_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_EQ_constant.npy", self.Sol_all)
				np.save(self.results_path + "vol_strain_3D_sub_EQ_constant.npy", self.vol_strain)
				np.save(self.results_path + "flow_velocity_3D_sub_EQ_constant.npy", self.flow_velocity)
				np.save(self.results_path + "psi_3D_EQ_mapped.npy", self.Psi)
				np.save(self.results_path + "sea_flux_3D_sub_EQ_constant.npy", self.sea_flux)
				np.save(self.results_path + "sea_flux_total_3D_sub_EQ_constant.npy", self.sea_flux_total)
			elif self.event == 'EQ':
				np.save(self.results_path + "Sol_surf_3D_EQ_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_EQ_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_EQ_constant.npy", self.Sol_all)
				np.save(self.results_path + "vol_strain_3D_EQ_constant.npy", self.vol_strain)
				np.save(self.results_path + "flow_velocity_3D_EQ_constant.npy", self.flow_velocity)
				np.save(self.results_path + "sea_flux_3D_EQ_constant.npy", self.sea_flux)
				np.save(self.results_path + "sea_flux_total_3D_EQ_constant.npy", self.sea_flux_total)

class Laplace3Dsteady:
	def __init__(self, data_file, origin, theta, ndim, mesh, boundaries, plot_figs, event, new_mesh, permeability,
	             sub_cycle_years, surface_k, ocean_k, loop, **kwargs):

		self.start = time.time()
		self.data_file = data_file
		self.origin = origin
		self.theta = theta
		self.mesh = mesh
		self.boundaries = boundaries
		self.plot_figs = plot_figs
		self.event = event
		self.new_mesh = new_mesh
		self.permeability = permeability
		self.sub_cycle_years = sub_cycle_years
		self.surface_k = surface_k
		self.ocean_k = ocean_k
		# solver = PETScKrylovSolver("gmres","petsc_amg")
		# solver = PETScKrylovSolver("cg","ilu")
		self.solver = LinearSolver('mumps')
		self.prm = self.solver.parameters
		# self.prm["form_compiler"]["representation"] = "uflacs"
		# self.prm['reuse_factorization'] = True  # Saves Factorization of LHS
		self.Qelement = FiniteElement('CG', self.mesh.ufl_cell(), 1)
		self.Q = FunctionSpace(self.mesh, self.Qelement)

		self.Kelement = FiniteElement('CG', self.mesh.ufl_cell(), 1)
		self.Kspace = FunctionSpace(self.mesh, self.Kelement)

		self.Q_gspace = VectorFunctionSpace(self.mesh, "DG", 0)
		self.Q_gspace1 = VectorFunctionSpace(self.mesh, "CG", 1)

		self.p = TrialFunction(self.Q)
		self.q = TestFunction(self.Q)
		self.ds = Measure("ds", domain=self.mesh, subdomain_data=self.boundaries)  #
		self.n = FacetNormal(self.mesh)  # normal vector
		self.pdim = self.Q.dim()
		self.geom_dim = ndim + 1

		##################### GPS STATION LOCATIONS ##############################

		self.x_gps = np.array([65308.5, 78441.2, 65707.9, 73555.6, 61257.8, 65291.6, \
		                       82112.3, 58804.1, 106822.8, 60152.4, 144960.9, 88660.1, \
		                       134829.2, 82409, 65214.4, 67249.7, 91916.7, 111195.9, 92905.4, \
		                       77176.5, 90809.3, 107806.1, 104439.5, 137051.8, 118095, 82238.4, \
		                       84123.2, 98197.5, 105163, 154799.8, 133655.9, 111467, 125498.1, \
		                       106995.1, 149450.7, 192061, 133903.2])

		self.y_gps = np.array([173604.8, 185529.2, 168017.6, 174095.8, 157386.9, 154352.5, \
		                       172814.3, 138685.6, 190893.8, 136426.2, 230041.9, 163931.9, \
		                       208246.3, 149993.8, 127836, 122418.9, 141284.3, 162013.4, \
		                       131283.2, 109747.4, 124059.5, 141615.9, 126126.3, 160541, \
		                       132800.4, 90851.5, 92291, 105309.4, 104713.8, 153672.9, \
		                       122823.1, 94557.3, 99698.5, 77971.8, 123239.8, 157764, 45232])

		self.z_gps = np.zeros(self.x_gps.shape)

		self.results_path = 'results/numpy_results/'
		self.var_path = 'results/numpy_variables/'
		self.paraviewpath = 'results/paraview/'
		if not os.path.exists(self.paraviewpath):
			os.makedirs(self.paraviewpath)

		self.extract_coords()
		print ("Coordinates extracted", (time.time() - self.start) / 60.0, "minutes")
		self.k_interpolation()
		print ("Permeability interpolated", (time.time() - self.start) / 60.0, "minutes")

		######################### POROELASTIC PARAMETERS #############################
		self.muf = Constant('1e-9')  # pore fluid viscosity [MPa s]
		self.d = self.p.geometric_dimension()  # number of space dimensions
		self.I = Identity(self.d)
		self.T = (self.I - outer(self.n, self.n))  # tangent operator for boundary condition
		self.year = 60 * 60 * 24 * 365

		################### INITIAL AND BOUNDARY CONDITIONS ######################
		####### Initial conditions: (('p0','ux0','uy0')) ##########
		self.w_init = Expression('0', degree=1)
		self.p0 = Function(self.Q)
		self.p0.interpolate(self.w_init)
		self.pb0 = Expression('0', degree=1)
		self.bcp1 = DirichletBC(self.Q, self.pb0, self.boundaries, 1)  # ocean surface (free flow condition)
		# self.bcp2 = DirichletBC(self.Q, ('1'), self.boundaries, 2)  # land surface (free flow condition)

		if self.event == 'topo':
			self.topo_flow_interp()
			self.bcp2 = DirichletBC(self.Q, self.p_topo, self.boundaries, 2)  # land surface
			self.bcs = [self.bcp1, self.bcp2]  # These are the BC's that are actually applied
		else:
			self.bcs = [self.bcp1]  # These are the BC's that are actually applied
		self.slab = 3  # the slab is labeled as boundary 3 for this mesh, but not always the case

		################### SET UP TO SAVE SOLUTION  ######################


		pfilename = self.paraviewpath + "pressure3D.pvd"
		p_velfilename = self.paraviewpath + "flux0_comp3D.pvd"
		p_velfilename1 = self.paraviewpath + "flux1_comp3D.pvd"

		self.pfile = File(pfilename)
		self.qfile = File(p_velfilename)
		self.qfile1 = File(p_velfilename1)

		print ("Solution matrices created", (time.time() - self.start) / 60.0, "minutes")

		################### SET UP, RUN, SOLVE and SAVE  ######################
		# self.plot_k()
		# print "Permeability mapped", (time.time() - self.start) / 60.0, "minutes"
		# self.plot_topo()
		# print "Topography mapped", (time.time() - self.start) / 60.0, "minutes"

		# quit()

		self.assemble_system()
		print ("System assembled", (time.time() - self.start) / 60.0, "minutes")
		self.solve_system()

	# self.save_solution()

	def extract_coords(self):
		if not os.path.exists(self.var_path):
			os.makedirs(self.var_path)
		if self.new_mesh == 'yes':  # create surface/GPS indices and save them

			self.x_all, self.y_all, self.z_all = SpatialCoordinate(self.mesh)

			xf = Function(self.Kspace)
			yf = Function(self.Kspace)
			zf = Function(self.Kspace)
			xm, ym, zm = SpatialCoordinate(self.mesh)

			self.xvec = xf.interpolate(xm).dat.data_ro
			self.yvec = yf.interpolate(ym).dat.data_ro
			self.zvec = zf.interpolate(zm).dat.data_ro


			bctest0 = DirichletBC(self.Kspace, (1), 1)  # ocean surface
			bctest1 = DirichletBC(self.Kspace, (2),  2)  # top surface
			#bctest2 = DirichletBC(self.Kspace, (1),  4)  # ocean wedge


			ptest = Function(self.Kspace)
			bctest0.apply(ptest.vector())
			bctest1.apply(ptest.vector())
			#bctest2.apply(ptest.vector())
			self.ocean_dofs = np.where(ptest.vector().array() == 1)[0]
			self.surface_dofs = np.where(ptest.vector().array() == 2)[0]

			np.save(self.var_path + "x_all_3D", self.xvec)
			np.save(self.var_path + "y_all_3D", self.yvec)
			np.save(self.var_path + "z_all_3D", self.zvec)
			np.save(self.var_path + "surface_dofs_3D", self.surface_dofs)
			np.save(self.var_path + "ocean_dofs_3D", self.ocean_dofs)



		elif self.new_mesh == 'no':  # load saved indices/variables
			self.x_all = np.load(self.var_path + "x_all_3D.npy")
			self.y_all = np.load(self.var_path + "y_all_3D.npy")
			self.z_all = np.load(self.var_path + "z_all_3D.npy")
			self.GPS_dofs = np.load(self.var_path + "GPS_dofs_3D.npy")
			self.surface_dofs = np.load(self.var_path + "surface_dofs_3D.npy")
			self.ocean_dofs = np.load(self.var_path + "ocean_dofs_3D.npy")
			self.size_p = np.load(self.var_path + "size_p_3D.npy")
			self.size_u = np.load(self.var_path + "size_u_3D.npy")

			self.ocean_dof_p = self.ocean_dofs[np.where(self.ocean_dofs < self.size_p)]
			self.ocean_dof_u = self.ocean_dofs[np.where(self.ocean_dofs >= self.size_p)] - self.size_p
			self.ocean_dof_pgrad = np.hstack((self.ocean_dof_p, (self.ocean_dof_p + self.ocean_dof_p.shape[0]),
			                                  (self.ocean_dof_p + 2 * self.ocean_dof_p.shape[0])))
			self.surface_dof_p = self.surface_dofs[np.where(self.surface_dofs < self.size_p)]
			self.surface_dof_u = self.surface_dofs[np.where(self.surface_dofs >= self.size_p)] - self.size_p
			self.surface_dof_pgrad = np.hstack((self.surface_dof_p, (self.surface_dof_p + self.surface_dof_p.shape[0]),
			                                    (self.surface_dof_p + 2 * self.surface_dof_p.shape[0])))

	def k_interpolation(self):
		log_kr = -18.0
		alpha_k = 0.9

		# self.coords_K = self.K.tabulate_dof_coordinates().reshape((self.Kspace.dim(), -1))
		bctest1 = DirichletBC(self.Kspace, (1.0), 1)  # ocean surface
		bctest2 = DirichletBC(self.Kspace, (1.0), 2)  # land surface
		#bctest3 = DirichletBC(self.Kspace, (1.0), 4)  # ocean surface near trench
		ktest = Function(self.Kspace)
		bctest1.apply(ktest)
		bctest2.apply(ktest)
		#bctest3.apply(ktest)
		self.dof_surfK = np.where(ktest.vector().array() == 1.0)[0]

		if self.permeability == 'mapped':

			# interpolate surface permeability onto all nodes
			kappa_surf = kappakriging(self.ocean_k, self.data_file, self.origin, self.theta,
			                          self.xvec, self.yvec, self.zvec, self.dof_surfK)
			z_surf = surfkriging(self.xvec, self.yvec, self.zvec, self.dof_surfK)

			print(self.zvec.min(), self.zvec.max())
			#quit()

			oceandof = np.where(z_surf < -10.)[0]
			kappa_surf[oceandof] = self.ocean_k
			self.kappa = Function(self.Kspace)
			self.kappa_vec = pow(10, (log_kr + ((kappa_surf - log_kr) *
			                                    pow((1 - (1e-3 * .5*(self.zvec - z_surf))), -alpha_k))))
			self.kappa.dat.data[:] = self.kappa_vec
			#self.kappa.dat.data[:] = 1e-12

		elif self.permeability == 'constant':
			# self.kappa = KExpression3DConstant(self.surface_k, self.ocean_k,
			#                                   log_kr, alpha_k, self.coords_K, self.dof_surfK, degree=1)
			kappa = Expression('pow(10, (log_kr + ((k_surf - log_kr) * pow((1 - (1e-3 * (x[2]))), -alpha_k))))',
			                   log_kr=log_kr, k_surf=self.surface_k, alpha_k=alpha_k)
			# kappa = Expression('pow(10, -10)')
			self.kappa = interpolate(kappa, self.Kspace)

	def topo_flow_interp(self):

		# self.coords_K = self.K.tabulate_dof_coordinates().reshape((self.Kspace.dim(), -1))

		bctest1 = DirichletBC(self.Kspace, (1.0), 2)  # land surface
		ktest = Function(self.Kspace)
		bctest1.apply(ktest)
		dof_surfK = np.where(ktest.vector().array() == 1.0)[0]

		wb = pyxl.load_workbook(self.data_file)
		topo_data = wb['topo_data3D']

		topo_p_surf = topokriging(topo_data, self.origin, self.theta,
		                          self.xvec, self.yvec, self.zvec, dof_surfK, self.var_path)

		# oceanz = np.where(self.zsurfall_p < 0.0)[0]
		low_p = np.where(topo_p_surf < 0.02)[0]
		# self.psurfall[oceanz] = 0.0
		topo_p_surf[low_p] = 0.02

		# NEW method
		self.p_topo = Function(self.Kspace)
		self.p_topo.dat.data[dof_surfK] = topo_p_surf
		#p_surf_func.vector()[self.dof_surfK] = topo_p_surf
		#self.p_topo = Expression(('ptopo'), ptopo=p_surf_func, degree=1)

	def plot_k(self):
		self.kappafile = File(self.paraviewpath + 'kappa2_3D.pvd')
		kappaplot = Function(self.Kspace)
		kappaplot.dat.data[:] = np.log10(self.kappa_vec)
		# kappaplot.interpolate(Expression('log10(kappa)', kappa=self.kappa, degree=1))
		# print((kappaplot.vector().min(), kappaplot.vector().max())
		self.kappafile.write(kappaplot)

	def plot_topo(self):
		self.topofile = File(self.paraviewpath + 'topo_3D.pvd')
		# print( kappaplot.vector().min(), kappaplot.vector().max()
		self.topofile.write(self.p_topo)

	def assemble_system(self):

		self.a_K = assemble(
			(self.kappa / self.muf) * inner(nabla_grad(self.p), nabla_grad(self.q)) * dx)  # Stiffness matrix

		self.A = self.a_K  # Left hand side (LHS)

	def streamfunction3D(self, constrained_domain=None):
		"""Stream function for a given 3D velocity field.
		The boundary conditions are weakly imposed through the term

			inner(q, grad(psi)*n)*ds,

		where u = curl(psi) is used to fill in for grad(psi) and set
		boundary conditions. This should work for walls, inlets, outlets, etc.
		"""

		vel = self.q_velocity
		qtest = TestFunction(self.Q)
		psi = TrialFunction(self.Q)
		a = inner(grad(qtest), grad(psi)) * dx - inner(qtest, dot(self.n, grad(psi))) * self.ds
		L = inner(grad(qtest), curl(vel)) * dx - dot(grad(qtest), cross(self.n, vel)) * self.ds
		psi_ = Function(self.Q)
		A = assemble(a)
		b = assemble(L)
		solver = LUSolver('petsc')
		b = assemble(L, tensor=b)
		solver.set_operator(A)
		solver.solve(psi_.vector(), b)

		return psi_

	def test_PSD(self):

		def is_pd(x):
			return np.all(np.linalg.eigvals(x) > 0)

		def is_psd(x, tol=1e-5):
			E, V = np.linalg.eigh(x)
			return np.all(E > -tol)

		def is_symmetric(x):
			return (x.transpose() == x).all()

		A = self.A.array()

		print('A is {}'.format('symmetric' if is_symmetric(A) else ('not symmetric')))

		np.linalg.cholesky(A)

	def solve_system(self):

		b = Vector()
		self.a_K.init_vector(b, 0)
		[bc.apply(b) for bc in self.bcs]
		[bc.apply(self.A) for bc in self.bcs]
		p = Function(self.Q)
		self.solver.set_operator(self.A)
		self.solver.solve(p.vector(), b)

		self.domain_flux = (-self.kappa / self.muf) * grad(p)  # element-wise flux vector
		# self.domain_flux = grad(p)  # element-wise flux vector
		self.bound_flux = assemble(inner(self.domain_flux, self.n) * self.ds(1))
		self.bound_flux2 = assemble(inner(self.domain_flux, self.n) * self.ds(2))
		self.bound_fluxall = assemble(inner(self.domain_flux, self.n) * self.ds)

		print ("outward flux through seafloor: ", self.bound_flux ) # * (3600*24*365.25)
		print ("outward flux through land surface: ", self.bound_flux2)
		print ("total boundary flux: ", self.bound_fluxall)

		q_comp = Function(self.Q_gspace)
		q_comp.assign(project(self.domain_flux, self.Q_gspace, solver_type='mumps'))
		q_comp1 = Function(self.Q_gspace1)
		q_comp1.assign(project(self.domain_flux, self.Q_gspace1, solver_type='mumps'))


		# self.psi = self.streamfunction3D()  # compute numerical streamfunction

		####################### SAVE SOLUTION ###########################
		p.rename('p', 'p')
		self.pfile << p
		q_comp.rename('q_comp', 'q_comp')
		self.qfile << q_comp
		q_comp1.rename('q_comp1', 'q_comp1')
		self.qfile1 << q_comp1

	def save_solution(self):
		if not os.path.exists(self.results_path):
			os.makedirs(self.results_path)
		if self.permeability == 'mapped':

			if self.event == 'topo':
				np.save(self.results_path + "Sol_surf_3D_topo_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_all_3D_topo_mapped.npy", self.Sol_all)
				np.save(self.results_path + "sea_flux_3D_topo_mapped.npy", self.sea_flux)
				np.save(self.var_path + "dtt_comsum_topo_mapped.npy", self.dtt_comsum)
				np.save(self.results_path + "flow_velocity_3D_topo_mapped.npy", self.flow_velocity)
				np.save(self.results_path + "vol_strain_3D_topo_mapped.npy", self.vol_strain)
				np.save(self.results_path + "Streamfun_3D_topo_mapped.npy", self.Psi)
				np.save(self.results_path + "sea_flux_total_3D_topo_mapped.npy", self.sea_flux_total)
		elif self.permeability == 'constant':
			if self.event == 'SSE':
				np.save(self.results_path + "Sol_surf_3D_SSE_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_SSE_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_SSE_constant.npy", self.Sol_all)
			elif self.event == 'sub':
				np.save(self.results_path + "Sol_surf_3D_sub_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_constant.npy", self.Sol_all)
			elif self.event == 'sub_EQ':
				np.save(self.results_path + "Sol_surf_3D_sub_EQ_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_EQ_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_EQ_constant.npy", self.Sol_all)
				np.save(self.results_path + "vol_strain_3D_sub_EQ_constant.npy", self.vol_strain)
				np.save(self.results_path + "flow_velocity_3D_sub_EQ_constant.npy", self.flow_velocity)
				np.save(self.results_path + "psi_3D_EQ_mapped.npy", self.Psi)
				np.save(self.results_path + "sea_flux_3D_sub_EQ_constant.npy", self.sea_flux)
				np.save(self.results_path + "sea_flux_total_3D_sub_EQ_constant.npy", self.sea_flux_total)
			elif self.event == 'EQ':
				np.save(self.results_path + "Sol_surf_3D_EQ_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_EQ_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_EQ_constant.npy", self.Sol_all)
				np.save(self.results_path + "vol_strain_3D_EQ_constant.npy", self.vol_strain)
				np.save(self.results_path + "flow_velocity_3D_EQ_constant.npy", self.flow_velocity)
				np.save(self.results_path + "sea_flux_3D_EQ_constant.npy", self.sea_flux)
				np.save(self.results_path + "sea_flux_total_3D_EQ_constant.npy", self.sea_flux_total)
